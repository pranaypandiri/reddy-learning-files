import pandas as pd
from io import BytesIO
import seaborn as sns
from collections import defaultdict
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import streamlit as st
import os
from langchain.memory import ConversationBufferMemory
from langchain.schema import HumanMessage, AIMessage, SystemMessage
from langchain_community.vectorstores import FAISS
from langchain_core.runnables import chain
from langchain_openai import AzureOpenAIEmbeddings
from langchain_core.documents import Document
from openai import AzureOpenAI as OpenAIAzureClient
from langchain_community.vectorstores import FAISS as EvalFAISS
from langchain_openai import AzureOpenAIEmbeddings as EvalEmbeddings
import re
from datetime import datetime
from gcsstorageunited import get_gcs_storage_manager

# Import our new modules
from user_simulation_manager_united import UserSimulationManager
from supervisor_dashboard_united import (
    render_user_management_tab,
    render_simulation_assignment_tab,
)


# Helper function to format eval scores consistently
def format_eval_score(score):
    """Format eval score to ensure consistent display without double %"""
    if isinstance(score, str):
        # Remove existing % if present and convert to float
        clean_score = score.replace("%", "").strip()
        try:
            return float(clean_score)
        except ValueError:
            return 0.0
    elif isinstance(score, (int, float)):
        return float(score)
    else:
        return 0.0


# Initialize the user simulation manager
user_sim_manager = UserSimulationManager()


def analyze_associate_behaviors(associate_data):
    """
    Analyze behavior data for an associate and categorize behaviors
    Args:
        associate_data: List of evaluation entries for the associate
    Returns:
        dict: Categorized behaviors with scores
    """
    behavior_totals = {}
    behavior_counts = {}
    behavior_max_totals = {}  # Track total max scores per behavior

    # Process each evaluation session
    for entry in associate_data:
        excel_bytes = entry.get("excel_bytes")
        if excel_bytes:
            try:
                with BytesIO(excel_bytes) as io:
                    wb = openpyxl.load_workbook(io, data_only=True)
                    if "Evaluation Table" in wb.sheetnames:
                        ws = wb["Evaluation Table"]
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            if len(row) >= 4:
                                behavior, _, max_score, score = row[:4]
                                try:
                                    score = float(score)
                                    max_score = float(max_score)
                                    if behavior:
                                        behavior_totals[behavior] = (
                                            behavior_totals.get(behavior, 0) + score
                                        )
                                        behavior_counts[behavior] = (
                                            behavior_counts.get(behavior, 0) + 1
                                        )
                                        behavior_max_totals[behavior] = (
                                            behavior_max_totals.get(behavior, 0)
                                            + max_score
                                        )
                                except (ValueError, TypeError):
                                    continue
            except Exception:
                continue

    if not behavior_totals:
        return {"weak": [], "moderate": [], "good": [], "overall_avg": 0}

    # Calculate average scores and average max scores for each behavior
    behavior_averages = {}
    behavior_max_averages = {}
    behavior_percentages = {}

    for behavior in behavior_totals.keys():
        avg_score = behavior_totals[behavior] / behavior_counts[behavior]
        avg_max_score = behavior_max_totals[behavior] / behavior_counts[behavior]
        percentage = (avg_score / avg_max_score) * 100 if avg_max_score > 0 else 0

        behavior_averages[behavior] = avg_score
        behavior_max_averages[behavior] = avg_max_score
        behavior_percentages[behavior] = percentage

    # Calculate overall average percentage across all behaviors
    overall_avg_percentage = sum(behavior_percentages.values()) / len(
        behavior_percentages
    )

    # Categorize behaviors based on absolute percentage thresholds
    weak_behaviors = []
    moderate_behaviors = []
    good_behaviors = []

    for behavior in behavior_averages.keys():
        percentage = behavior_percentages[behavior]
        avg_score = behavior_averages[behavior]
        avg_max = behavior_max_averages[behavior]

        if percentage < 60.0:  # Below 60% is weak performance
            weak_behaviors.append((behavior, avg_score, avg_max, percentage))
        elif percentage >= 80.0:  # 80% and above is good performance
            good_behaviors.append((behavior, avg_score, avg_max, percentage))
        else:  # 60% to 79% is moderate (needs improvement)
            moderate_behaviors.append((behavior, avg_score, avg_max, percentage))

    # Sort by percentages (descending for good, ascending for weak)
    weak_behaviors.sort(key=lambda x: x[3])  # Sort by percentage
    moderate_behaviors.sort(key=lambda x: x[3])
    good_behaviors.sort(key=lambda x: x[3], reverse=True)

    return {
        "weak": weak_behaviors,
        "moderate": moderate_behaviors,
        "good": good_behaviors,
        "overall_avg": overall_avg_percentage,
        "behavior_averages": behavior_averages,
        "behavior_percentages": behavior_percentages,
    }


def create_behaviors_graph(evaluation_data, title_prefix=""):
    """
    Create a behaviors graph from evaluation data
    Args:
        evaluation_data: List of evaluation entries with excel_bytes
        title_prefix: Optional prefix for the graph title
    """
    behavior_totals = {}
    behavior_counts = {}
    total_sessions = len(evaluation_data)

    for entry in evaluation_data:
        eb = entry.get("excel_bytes")
        if eb:
            try:
                with BytesIO(eb) as io:
                    wb = openpyxl.load_workbook(io, data_only=True)
                    if "Evaluation Table" in wb.sheetnames:
                        ws = wb["Evaluation Table"]
                        # Process data
                        for r in ws.iter_rows(min_row=2, values_only=True):
                            b, _, mx, sc, *_ = r
                            try:
                                sc = float(sc)
                                if b:
                                    behavior_totals[b] = behavior_totals.get(b, 0) + sc
                                    behavior_counts[b] = behavior_counts.get(b, 0) + 1
                            except Exception:
                                continue
            except Exception:
                continue

    # Display session summary before the graph
    st.info(
        f"📊 **Analysis Summary:** {total_sessions} total training sessions completed"
    )

    if behavior_totals:
        behaviors = list(behavior_totals.keys())
        # Calculate averages instead of totals
        avg_scores = [behavior_totals[b] / behavior_counts[b] for b in behaviors]
        session_counts = [behavior_counts[b] for b in behaviors]

        plot_df = pd.DataFrame(
            {
                "Behavior": behaviors,
                "Average Score": avg_scores,
                "Sessions": session_counts,
            }
        )

        # Sort by average score for better visualization
        plot_df = plot_df.sort_values("Average Score", ascending=True)

        # Set the style
        sns.set_style("whitegrid")
        plt.style.use("seaborn-v0_8")

        # Calculate dynamic figure size based on number of behaviors
        fig_height = max(6, len(behaviors) * 0.6)
        fig_width = 12

        fig, ax = plt.subplots(figsize=(fig_width, fig_height))

        # Create a beautiful horizontal bar plot
        bars = sns.barplot(
            data=plot_df,
            y="Behavior",
            x="Average Score",
            ax=ax,
            palette="viridis",
            orient="h",
        )

        # Add value labels on the bars with session count
        for i, (behavior, avg_score, session_count) in enumerate(
            zip(plot_df["Behavior"], plot_df["Average Score"], plot_df["Sessions"])
        ):
            ax.text(
                avg_score + max(avg_scores) * 0.01,
                i,
                f"{avg_score:.1f} ({session_count} sessions)",
                va="center",
                ha="left",
                fontsize=10,
                fontweight="bold",
            )

        # Customize the plot
        ax.set_xlabel("Average Score", fontsize=12, fontweight="bold")
        ax.set_ylabel("Behavior", fontsize=12, fontweight="bold")
        title = (
            f"{title_prefix}Average Behavior Performance Scores"
            if title_prefix
            else "Average Behavior Performance Scores"
        )
        ax.set_title(
            title,
            fontsize=14,
            fontweight="bold",
            pad=20,
        )

        # Improve the y-axis labels (behavior names)
        ax.tick_params(axis="y", labelsize=10)
        ax.tick_params(axis="x", labelsize=10)

        # Set x-axis limits with some padding
        ax.set_xlim(0, max(avg_scores) * 1.15)

        # Customize grid
        ax.grid(True, linestyle="--", alpha=0.7, color="gray", axis="x")
        ax.set_axisbelow(True)

        # Add a subtle background color gradient
        for i, bar in enumerate(bars.patches):
            bar.set_edgecolor("white")
            bar.set_linewidth(1)

        plt.tight_layout()
        st.pyplot(fig)
        plt.close()
    else:
        st.info("No behavior scores to plot.")


# --- SUPERVISOR ASSIGNMENTS CACHING ---
@st.cache_data
def cached_supervisor_assignments(username):
    """Return supervisor assignments for a specific user (cached)."""
    if not username:
        return []
    try:
        assignments = user_sim_manager.get_user_simulation_assignments(username)
        return [assignment["simulation"] for assignment in assignments]
    except Exception:
        return []


# --- POLICY DOCUMENTS FAISS DB Caching ---
@st.cache_resource
def get_policy_faiss_db():
    """Load the policy documents vector store created by Embedding_united.py"""
    embeddings = CustomerSupportTrainer.OAI_embedder()
    try:
        policy_db = FAISS.load_local(
            "faiss_store_united_policy_docs",
            embeddings,
            allow_dangerous_deserialization=True,
        )

        # Debug: Show available themes in the policy database
        available_themes = set()
        for doc_id, doc in policy_db.docstore._dict.items():
            theme = doc.metadata.get("theme", "Unknown")
            available_themes.add(theme)
        print(
            f"[DEBUG] Available themes in United policy database: {sorted(available_themes)}"
        )

        return policy_db
    except Exception as e:
        print(f"[ERROR] Could not load United policy documents vector store: {e}")
        st.error(f"Failed to load United policy documents: {e}")
        return None


def get_available_policy_themes():
    """Get all available themes from the United policy documents vector store"""
    try:
        policy_db = get_policy_faiss_db()
        available_themes = set()
        for doc_id, doc in policy_db.docstore._dict.items():
            theme = doc.metadata.get("theme", "Unknown")
            if theme != "Unknown":
                available_themes.add(theme)
        return sorted(list(available_themes))
    except Exception as e:
        print(f"[ERROR] Could not get available themes: {e}")
        return [
            "buy miles",
            "earn credit",
            "account changes",
            "premier status",
            "missing credit",
            "upgrades",
            "use miles",
        ]


# Azure OpenAI Configuration
client = OpenAIAzureClient(
    api_key="92dc252cdb0c4079b4712a9ead4179ca",
    api_version="2024-12-01-preview",
    azure_endpoint="https://azureaitest4641590782.openai.azure.com/",
)

os.environ["OPENAI_API_KEY"] = "92dc252cdb0c4079b4712a9ead4179ca"


class CustomerSupportTrainer:
    GOODBYE_KEYWORDS = [
        "bye",
        "goodbye",
        "have a nice day",
        "alright, thanks for your help.",
        "resolved",
        "issue fixed",
        "problem solved",
        "Thanks again!",
    ]

    def __init__(self):
        # Use buffer memory for context management
        self.memory = ConversationBufferMemory(
            memory_key="chat_history", return_messages=True
        )
        # Static first message - Banking fraud scenario
        # self.initial_customer_message = """
        # Hi, I need help. After the latest Windows update, my store computer has become very slow and it's affecting checkout. Can you help me fix this?
        # """
        # Conversation started flag
        self.phase_switched = False
        self.conversation_started = False

    @staticmethod
    def OAI_embedder():
        embedding_model = AzureOpenAIEmbeddings(
            model="text-embedding-3-large",
            api_key="83265f364fe844f298b2d4f8a5a39426",
            openai_api_type="azure",
            azure_endpoint="https://genai-demos.openai.azure.com/",
            api_version="2024-02-15-preview",
        )
        return embedding_model

    # Add this method to CustomerSupportTrainer class

    def detect_assistance_offer(self, support_rep_message):
        # Get conversation history to check if this is the very first message
        memory_vars = self.memory.load_memory_variables({})
        chat_history = memory_vars.get("chat_history", [])

        # If this is the first message in conversation, never trigger phase switch
        if len(chat_history) == 0:
            return False

        if self.phase_switched:
            return True

        assistance_triggers = [
            "anything else",
            "any other",
            "anything more",
            "further assistance",
            "help you with anything else",
            "other questions",
            "additional help",
            "something else",
            "else i can help",
            "anything additional",
            "other concerns",
            "more assistance",
            "help with anything",
            "other issues",
        ]

        message_lower = support_rep_message.lower()
        for trigger in assistance_triggers:
            if trigger in message_lower:
                self.phase_switched = True
                return True
        return False

    def gen_query_cont(
        self, support_rep_message, examples, customer_tone, theme, all_matching_docs
    ):
        """
        Generate policy-based customer questions when support rep offers additional assistance
        Uses random sampling of 4 policy chunks to create scenarios, stored in session to avoid re-creation
        """
        try:
            # Get conversation history up to this point
            memory_vars = self.memory.load_memory_variables({})
            chat_history = memory_vars.get("chat_history", [])

            # Build conversation history text for context
            history_text = ""
            for message in chat_history:
                if isinstance(message, HumanMessage):
                    history_text += f"Support Rep: {message.content}\n"
                elif isinstance(message, AIMessage):
                    history_text += f"Customer: {message.content}\n"

            # Check if we already have a policy scenario stored in session for this switch
            if "policy_switch_scenario" not in st.session_state:
                # Random sampling of 4 policy chunks
                if all_matching_docs and len(all_matching_docs) >= 4:
                    import random

                    selected_chunks = random.sample(all_matching_docs, 4)
                elif all_matching_docs:
                    selected_chunks = all_matching_docs  # Use all if less than 4
                else:
                    return "I think that covers everything I needed to know. Thank you for your help!"

                # Create policy context from selected chunks
                policy_context = ""
                for i, doc in enumerate(selected_chunks, 1):
                    policy_intent = doc.metadata.get("policy_intent", "General policy")
                    chunk_id = doc.metadata.get("chunk_id", f"chunk_{i}")
                    policy_context += f"POLICY CHUNK {i} - {policy_intent}:\n"
                    policy_context += f"Chunk ID: {chunk_id}\n"
                    policy_context += f"Content: {doc.page_content}\n"
                    policy_context += "=" * 60 + "\n"

                # Generate scenario based on selected policy chunks
                scenario_prompt = f"""
You are tasked with creating a realistic customer question scenario based on company policy documents.

CONTEXT:
- Theme: {theme}
- Customer has been offered additional assistance by support rep
- The conversation history shows what has already been discussed and resolved

EXISTING CONVERSATION HISTORY:
{history_text}

AVAILABLE POLICY CHUNKS FOR NEW SCENARIO:
{policy_context}

INSTRUCTIONS:
1. Create ONE realistic customer question/scenario based on the policy chunks above
2. The question must be DIFFERENT from what's already been discussed in the conversation history
3. Base the question on specific policy situations, rules, or procedures from the chunks
4. Make it something a real customer would ask about after being offered additional help
5. Keep it focused on the theme domain
6. Don't repeat or rephrase anything already resolved in the conversation

Generate ONLY the customer scenario/question (1-2 sentences max) that would naturally come up based on the policies.
"""

                # Call LLM to generate scenario
                response = client.chat.completions.create(
                    model="gpt-4.1",
                    messages=[
                        {
                            "role": "system",
                            "content": "You are an expert at creating realistic customer service scenarios based on company policies.",
                        },
                        {"role": "user", "content": scenario_prompt},
                    ],
                    max_tokens=150,
                    temperature=0.5,
                )

                generated_scenario = response.choices[0].message.content.strip()

                # Store in session state to avoid regeneration on streamlit reruns
                st.session_state.policy_switch_scenario = {
                    "scenario": generated_scenario,
                    "selected_chunks": selected_chunks,
                    "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                }

                # Also update current policy session similar to gen_cus_policy_response for evaluation
                chunk_mapping = {}
                for i, doc in enumerate(selected_chunks, 1):
                    chunk_id = doc.metadata.get("chunk_id", f"CHUNK_{i}")
                    chunk_mapping[1] = chunk_mapping.get(1, "") + f"{chunk_id}, "

                # Clean up the chunk mapping (remove trailing comma)
                if 1 in chunk_mapping:
                    chunk_mapping[1] = chunk_mapping[1].rstrip(", ")

                st.session_state.current_policy_session = {
                    "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "department": "Policy Inquiry",
                    "theme": theme,
                    "selected_policies": selected_chunks,
                    "scenarios": [generated_scenario],
                    "scenario_chunks": chunk_mapping,
                }

            else:
                # Use existing scenario from session state
                generated_scenario = st.session_state.policy_switch_scenario["scenario"]
                selected_chunks = st.session_state.policy_switch_scenario[
                    "selected_chunks"
                ]

            # Now generate customer response using the scenario and current support rep message
            policy_context_for_response = ""
            for i, doc in enumerate(selected_chunks, 1):
                policy_intent = doc.metadata.get("policy_intent", "General policy")
                chunk_id = doc.metadata.get("chunk_id", f"chunk_{i}")
                policy_context_for_response += f"POLICY CHUNK {i} - {policy_intent}:\n"
                policy_context_for_response += f"Chunk ID: {chunk_id}\n"
                policy_context_for_response += f"Content: {doc.page_content}\n"
                policy_context_for_response += "=" * 60 + "\n"

            # Generate actual customer response
            customer_prompt = f"""
You are an AI Assistant mimicking a customer in a support chat. Your tone should match the specified customer tone: {customer_tone}.

SCENARIO CONTEXT:
- Theme: {theme}
- Customer Tone: {customer_tone}
- Current Scenario: {generated_scenario}

CONVERSATION HISTORY:
{history_text}

SUPPORT REP JUST SAID: {support_rep_message}

POLICY CONTEXT FOR RESOLUTION ANALYSIS:
{policy_context_for_response}

Generate only the customer's next message (1-2 short sentences) that fits the scenario and tone.
"""

            response = client.chat.completions.create(
                model="gpt-4.1",
                messages=[
                    {
                        "role": "system",
                        "content": "You are a customer asking about company policies after being offered additional help.",
                    },
                    {"role": "user", "content": customer_prompt},
                ],
                max_tokens=100,
                temperature=0.4,
            )

            customer_response = response.choices[0].message.content.strip()

            return customer_response

        except Exception as e:
            print(f"[ERROR] Error in gen_query_cont: {str(e)}")
            return f"Actually, I think I'm all set. Thank you for your help!"

    def reset_conversation(self):
        """Reset conversation memory and policy scenarios"""
        # Clear LangChain memory
        self.memory.clear()

        # Reset all instance flags to initial state
        self.conversation_started = False
        self.phase_switched = False

        # Clear all policy-related session state variables
        if "policy_scenarios" in st.session_state:
            st.session_state.policy_scenarios = []
        if "current_policy_session" in st.session_state:
            st.session_state.current_policy_session = None
        if "policy_switch_scenario" in st.session_state:
            del st.session_state.policy_switch_scenario

        # Additional safety check - clear any other potential session state variables
        session_keys_to_clear = [
            "selected_policies",
            "scenario_chunks",
            "policy_context",
            "current_scenario",
        ]

        for key in session_keys_to_clear:
            if key in st.session_state:
                del st.session_state[key]

        print("✅ Conversation reset complete")
        print(
            f"[DEBUG]   - current_policy_session: {'CLEARED' if 'current_policy_session' not in st.session_state or st.session_state.current_policy_session is None else 'NOT CLEARED'}"
        )
        print(
            f"[DEBUG]   - policy_switch_scenario: {'CLEARED' if 'policy_switch_scenario' not in st.session_state else 'NOT CLEARED'}"
        )

    def get_conversation_history(self):
        """Get formatted conversation history"""
        memory_vars = self.memory.load_memory_variables({})
        return memory_vars.get("chat_history", [])

    def debug_memory(self):
        """Debug method to check if summarization is happening"""
        # Check if memory has a moving_summary_buffer attribute
        if (
            hasattr(self.memory, "moving_summary_buffer")
            and self.memory.moving_summary_buffer
        ):
            return f"SUMMARY CREATED: {self.memory.moving_summary_buffer}"
        else:
            return "No summary yet"

    def has_conversation_ended(self, chat_history, support_rep_message):
        """Check if the conversation has ended based on goodbye keywords"""
        # Check if support rep message contains goodbye keywords
        if any(kw in support_rep_message.lower() for kw in self.GOODBYE_KEYWORDS):
            return True

        # Check if the last customer message contained goodbye keywords
        if chat_history:
            last_message = chat_history[-1]
            if isinstance(last_message, AIMessage):  # Customer's last message
                if any(
                    kw in last_message.content.lower() for kw in self.GOODBYE_KEYWORDS
                ):
                    return True

        return False

    def gen_cus_policy_response(
        self, support_rep_message, policy_docs, customer_tone, department, theme
    ):

        try:
            # Get conversation history
            memory_vars = self.memory.load_memory_variables({})
            chat_history = memory_vars.get("chat_history", [])

            # Modular goodbye detection
            if self.has_conversation_ended(chat_history, support_rep_message):
                # Conversation ended - skip generation
                self.memory.save_context({"input": support_rep_message}, {"output": ""})
                self.conversation_started = False
                return None

            # Initialize session tracking for policy scenarios if not exists
            if "policy_scenarios" not in st.session_state:
                st.session_state.policy_scenarios = []
            if "current_policy_session" not in st.session_state:
                st.session_state.current_policy_session = None

            # Check if this is the first customer interaction (no scenarios exist yet)
            # OR if we need to create scenarios for a new department/theme combination
            need_new_scenarios = (
                len(st.session_state.policy_scenarios) == 0
                or st.session_state.current_policy_session is None
                or st.session_state.current_policy_session.get("department")
                != department
                or st.session_state.current_policy_session.get("theme") != theme
            )

            selected_policies = []
            scenarios = []

            if need_new_scenarios:
                # Creating new scenarios (logging reduced)

                # Select 4 random policy documents from available ones
                if policy_docs and len(policy_docs) > 0:
                    import random

                    # Pick up to 4 policies (or all if less than 4)
                    num_to_select = min(4, len(policy_docs))
                    selected_policies = random.sample(policy_docs, num_to_select)

                    # Selected policies logged (reduced verbosity)

                    # Create 2 scenario-based queries from the selected policies
                    scenarios, scenario_chunks = self.create_policy_scenarios(
                        selected_policies, department, theme
                    )

                    # Create new policy session entry
                    st.session_state.current_policy_session = {
                        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "department": department,
                        "theme": theme,
                        "selected_policies": selected_policies,  # Store full policy documents
                        "scenarios": scenarios,
                        "scenario_chunks": scenario_chunks,  # Store chunk mappings
                    }

                    # Also store in the policy_scenarios list for tracking
                    scenario_entry = {
                        "timestamp": st.session_state.current_policy_session[
                            "timestamp"
                        ],
                        "department": department,
                        "theme": theme,
                        "selected_policies": [
                            {
                                "content": doc.page_content[:200]
                                + "...",  # Truncate for storage
                                "theme": doc.metadata.get("theme", ""),
                                "policy_intent": doc.metadata.get("policy_intent", ""),
                            }
                            for doc in selected_policies
                        ],
                        "scenarios": scenarios,
                        "scenario_chunks": scenario_chunks,  # Store chunk mappings
                    }
                    st.session_state.policy_scenarios.append(scenario_entry)

                    # Created and stored scenarios (logging reduced)
            else:
                # Reuse existing scenarios and policies from current session
                selected_policies = st.session_state.current_policy_session[
                    "selected_policies"
                ]
                scenarios = st.session_state.current_policy_session["scenarios"]
                scenario_chunks = st.session_state.current_policy_session.get(
                    "scenario_chunks", {}
                )

                # Log only the count of reused policy chunks
                print(f"� Reusing {len(selected_policies)} existing policy chunks")

            # Build conversation history for context
            history_text = ""
            for message in chat_history:
                if isinstance(message, HumanMessage):
                    history_text += f"Support Rep: {message.content}\n"
                elif isinstance(message, AIMessage):
                    history_text += f"Customer: {message.content}\n"

            # Create policy context from selected documents
            policy_context = ""
            if selected_policies:
                policy_context = "RELEVANT POLICY INFORMATION:\n"
                policy_context += "=" * 60 + "\n"

                # Get all required chunks from scenario_chunks
                required_chunks = []
                if "scenario_chunks" in st.session_state.current_policy_session:
                    scenario_chunks = st.session_state.current_policy_session[
                        "scenario_chunks"
                    ]
                    for chunk_list in scenario_chunks.values():
                        # Handle cases like "CHUNK_24" or "CHUNK_23, CHUNK_25"
                        chunks = [chunk.strip() for chunk in chunk_list.split(",")]
                        required_chunks.extend(chunks)

                print(f"[DEBUG] Required chunks from scenarios: {required_chunks}")

                # Normalize required chunks to handle case variations (CHUNK_X vs chunk_x)
                normalized_required_chunks = set()

                for chunk in required_chunks:
                    # Extract the number from patterns like "CHUNK_18" or "chunk_18"
                    chunk_num = (
                        chunk.replace("CHUNK_", "").replace("chunk_", "").strip()
                    )
                    normalized_required_chunks.add(chunk_num)

                print(
                    f"[DEBUG] Normalized required chunk numbers: {sorted(normalized_required_chunks)}"
                )

                for i, doc in enumerate(selected_policies, 1):
                    doc_chunk_id = doc.metadata.get("chunk_id", f"CHUNK_{i}")

                    # Extract chunk number from document chunk_id (e.g., "chunk_10" -> "10")
                    doc_chunk_num = (
                        doc_chunk_id.replace("CHUNK_", "").replace("chunk_", "").strip()
                    )

                    # Check if this document's chunk number is in the required chunks
                    if doc_chunk_num in normalized_required_chunks:
                        policy_intent = doc.metadata.get(
                            "policy_intent", "General policy"
                        )
                        policy_context += (
                            f"POLICY CHUNK {i} - {policy_intent.upper()}:\n"
                        )
                        policy_context += f"Chunk ID: {doc_chunk_id}\n"
                        policy_context += f"{doc.page_content}\n"
                        policy_context += "=" * 60 + "\n"
                    else:
                        print(
                            f"[DEBUG] ❌ Skipped policy chunk {doc_chunk_id} - not in required chunks"
                        )

                # 🌟 POLICY CONTEXT LOGGING 🌟
                # Essential policy context logging
                print(f"📋 Using {len(selected_policies)} policy chunks for response")

            # Get the single scenario (no need for indexing since we only have one)
            current_scenario = ""
            current_scenario_chunks = ""
            if scenarios and len(scenarios) > 0:
                current_scenario = scenarios[
                    0
                ]  # Always use the first (and only) scenario
                # Get chunk information for the scenario
                if "scenario_chunks" in st.session_state.current_policy_session:
                    scenario_chunks = st.session_state.current_policy_session[
                        "scenario_chunks"
                    ]
                    current_scenario_chunks = scenario_chunks.get(1, "Not specified")

                print(f"[DEBUG] Using scenario: {current_scenario}")
                print(f"[DEBUG] Required chunks: {current_scenario_chunks}")

            # Create customer response prompt with scenario and policy context
            customer_prompt = f"""
            You are an AI Assistant mimicking a **customer** in a support chat. Your tone should match the specified customer tone: {customer_tone}.

            SCENARIO CONTEXT:
            - Department: {department}
            - Theme: {theme}
            - Customer Tone: {customer_tone}
            - Current Scenario: {current_scenario}
            - Required Policy Chunks: {current_scenario_chunks}

           

            SUPPORT REP JUST SAID: {support_rep_message}

            POLICY CONTEXT FOR RESOLUTION ANALYSIS:
            {policy_context}

            YOUR TASK (strictly follow all points):

            **QUESTION TRACKING INSTRUCTIONS (MOST IMPORTANT):**
            1. **Look at the CONVERSATION HISTORY  {history_text}:**
               - What questions you already asked (recognize if the same question was rephrased)
               - What responses the rep gave (use policy context to assess if good or bad)
               - How many times you asked about the same TOPIC (not just exact wording)
               - To check the attempts for escalation if needed

            **Critical**: if the same variation of the question was present in the {history_text}, do NOT ask it again.

            **Critical**: if Both scenario questions were already asked, assess the quality of answers received with the above POLICY CONTEXT FOR RESOLUTION ANALYSIS section:
 :
               - **If got GOOD answers**: "I think that covers everything I needed to know. Thank you for your help!"
               - **If got BAD answers to both questions**: Express dissatisfaction before ending:
                 • "I'm not fully satisfied with the answers, but I'll try calling back later. Thanks anyway."
                 • "I didn't get clear answers to my concerns, but I guess that's all for now. Thank you."
                 • "I'm still not clear on these issues, but I'll figure it out another way. Thanks for trying."

            2. **Question Topic Recognition:**
               - Count attempts by TOPIC, not exact wording
               - Examples of same topic different wording:
                 • "Can children sit in exit rows?" = "Are kids allowed in exit row seats?" = "What about minors in emergency exit seats?"
                 • "What's the pet fee?" = "How much for bringing my dog?" = "Cost for animal travel?"
               - **IMPORTANT**: If you asked about a topic twice (even with different words), move to next question

            3. **Answer Quality Assessment (Use Policy Context):**
               - **MANDATORY**: Compare rep's answer against POLICY CONTEXT FOR RESOLUTION ANALYSIS
               - **Good answer**: Directly matches or aligns with policy information provided
               - **Bad answer**: Contradicts, differs from, or ignores the policy information shown above
               - **"I don't know" answer**: Rep explicitly admits they don't know
               - **CRITICAL**: You MUST assess EVERY answer against the policy context - do not skip this step

            4. **Question Progression Logic (Your Exact Flow):**
               - **First attempt**: Ask question from scenario
               - **Got bad/wrong answer (based on policy context)**: ALWAYS express confusion and ask to check again
                 Example tones: 
                 • "That doesn't sound right based on what I've heard. Can you check once more?"
                 • "I'm not sure about that answer. Could you double-check the policy?"
                 • "Hmm, that seems incorrect. Can you verify that for me?"
               - **Second attempt with bad answer**: Politely move to next question
                 Example tones:
                 • "Okay, I'm not getting a clear answer on that. Let me ask about something else..."
                 • "Alright, let's move on. I have another question about..."
                 • "I'll leave that for now. Can you help me with..."
               - **If all questions asked**: End conversation politely
               
               **MANDATORY RULE**: Before moving to next question, you MUST check if the rep's answer contradicts the policy context. If it does, challenge it first.

            5. **Escalation Logic (Only for "I don't know" responses):**
               - **Only escalate if rep explicitly says**: "I don't know", "I'm not sure", "I can't help with that"
               - **Escalation request**: "Could you please connect me with a supervisor?"
               - **If escalation denied**: "I'm disappointed this can't be resolved, but thanks for your time."

            **NEVER ask about the same TOPIC more than twice, even if you rephrase the question.**

            CRITICAL INSTRUCTIONS:
            1. **MULTI-QUESTION SCENARIO FOCUS**: You are working on this ONE scenario with multiple questions: "{current_scenario}"
                - This scenario contains multiple related questions - address them one by one
                - Do NOT ask multiple questions in one response
                - Move through questions in order based on chat history tracking
            2. **Don't create new scenarios or queries based on the policy context you have - stick to the scenario: {current_scenario}.**
            3. **Use the given policy context only as background awareness, not to challenge or correct the support rep’s answers.
            4. **If the rep asks about the policy, respond briefly (2–3 sentences) to acknowledge it, but don't argue.
            5. **When your tone is angry, express your dissatisfaction briefly at the start, but do not repeat or exaggerate anger in every sentence.

            Guide for your response:
            1. **If the Support Rep's message asks for patience or says they are checking/waiting/processing (e.g., "I'll check," "wait please," "one moment"),** do NOT move the scenario forward or repeat any earlier query. Instead, respond naturally and briefly with patience or acknowledgment (e.g., "Sure, take your time.", "Thanks, I'll wait for your update.").

            2. **If asked for personal details (such as a MileagePlus account number), reply with a plausible, randomly-generated 8–10 digit number (e.g., "My account number is 387459263.")**. Generate a new realistic number every time unless you have already shared one earlier in this conversation (then re-use it if asked again).

            3. **Ask just one new question or make one comment per turn related only to the current policy scenario — never more than one. Make all responses concise (1–2 sentences max) and do NOT repeat previous queries unless specifically requested by the Support Rep. Avoid re-phrasing or restating the same question if it's already pending.**

            4. **Response Acknowledgment (Based on Current Approach):**
               - **If rep gives a GOOD answer**: Acknowledge briefly and move to next question:
                 • "Thank you, that helps. Now I have another question..."
                 • "Got it, thanks. I also wanted to ask about..."
                 • "That makes sense. Can you also help me with..."
               - **If rep gives a BAD answer**: Follow your 2-attempt rule (express confusion, ask again, then move on)
               - **If rep says "I don't know"**: Ask for escalation to supervisor
               - **If rep says goodbye**: End conversation politely

            5. **Humanic Question Transitions (After 2 attempts OR good answer):**
                - Connect smoothly when moving between questions with natural transitions:
                - Example: "Thank you for that. I have another question about [next topic]…"
                - Example: "Alright, moving on. Can you help me with [next topic]…"
                - Example: "Got it. I also wanted to ask about [next topic]…"
                - Example: "Okay, let me ask about something else. [next topic]…"
              - Do NOT abruptly switch topics - always use smooth, human-like transitions

            6. **NATURAL HUMAN-LIKE BEHAVIOR INSTRUCTIONS:**
               - **Provide additional context naturally**: When asking questions, occasionally mention relevant personal details that a real customer would share:
                 • "I'm planning a trip next month and..."
                 • "My family and I are traveling and..."
                 • "I just joined MileagePlus recently and..."
                 • "I'm trying to book a flight and..."
               
               - **Show genuine curiosity**: Ask follow-up questions that show you're engaged:
                 • "That's interesting, does that apply to all flights?"
                 • "Good to know! Is there anything else I should be aware of?"
                 • "Thanks! Just to be sure, does that include..."
               
               - **Express natural reactions**: React like a real person would:
                 • "Oh wow, I didn't know that!"
                 • "That's really helpful, thank you!"
                 • "Hmm, let me think about that..."
                 • "Perfect, that's exactly what I needed to know!"
               
               - **Volunteer relevant information**: Share details that help the conversation flow:
                 • If asking about miles, mention "I took a flight last week"
                 • If asking about policies, mention your specific situation
                 • Provide context that makes your questions more realistic
               
               - **Be conversational, not transactional**: Don't just ask questions mechanically - engage in natural dialogue while staying focused on your scenario
              - Keep conversations flowing naturally with polite connectors

            6. **Style and Etiquette:**
               - Do not invent unrelated hypothetical problems.
               - Use polite, natural, and human-like language. Vary your interjections subtly ("Thanks!", "I appreciate it.", "Alright, I'll wait.").
               - Always maintain the specified customer tone, especially if told to be "angry" or "confused": express those emotions briefly once at the start of the query, but never repeat or exaggerate in each sentence.

            6. **Important Reminders:**
               - NEVER repeat the same question more than twice
               - If rep says "wait" or "checking", respond with patience: "Sure, take your time"
               - Follow the 2-attempt rule: Ask → Express confusion → Ask again → Move to next question


            -----
            **Generate only the customer's next message based on all of the above.**
            Keep your reply natural, concise, and maximally relevant to the current scenario and flow of the conversation.
            """

            response = client.chat.completions.create(
                model="gpt-4.1",
                messages=[
                    {
                        "role": "system",
                        "content": "You are a customer with a specific concern asking for help in a support chat.",
                    },
                    {"role": "user", "content": customer_prompt},
                ],
                max_tokens=200,
                temperature=1.0,
                top_p=0.7,
            )

            customer_response = response.choices[0].message.content.strip()

            # Check if customer response contains goodbye/closing
            if any(kw in customer_response.lower() for kw in self.GOODBYE_KEYWORDS):
                self.conversation_started = False

            # Save to memory
            self.memory.save_context(
                {"input": support_rep_message}, {"output": customer_response}
            )

            return customer_response

        except Exception as e:
            print(f"[ERROR] Error in gen_cus_policy_response: {str(e)}")
            return f"Error generating response: {str(e)}"

    def create_policy_scenarios(self, selected_policies, department, theme):
        """
        Create 1 multi-question scenario from the selected policy documents
        """
        try:
            print(f"🎬 Generating scenarios for {department} - {theme}")
            policy_summaries = []
            for i, doc in enumerate(selected_policies, 1):
                policy_intent = doc.metadata.get("policy_intent", "General policy")
                content_summary = doc.page_content
                policy_id = doc.metadata.get("chunk_id", f"Policy-{i}")
                print(f"  Chunk {i}: {policy_id} - {content_summary}")

                policy_summaries.append(
                    f"Policy Intent: {policy_intent}\nContent: {content_summary} \nID: {policy_id}"
                )

            combined_policies = "\n\n".join(policy_summaries)

            scenario_prompt = f"""
            Based on the following company policy documents, create 1 realistic customer service scenario
            that would require knowledge of these policies to resolve properly.

            Department: {department}
            Theme: {theme}

            POLICY DOCUMENTS:
            {combined_policies}

            Generate 1 customer scenario that includes at least 2 related questions.
            The scenario should:
            1. Be a realistic customer problem with multiple related questions
            2. Require multiple policy documents from the provided policies to resolve
            3. Be specific to the {department} department and {theme} theme
            4. Include at least 2 questions that a customer would naturally ask in sequence
            5. **IMPORTANT: Mention which CHUNK numbers from ALL available chunks are needed to answer the scenario**


            Format EXACTLY like this:

            Scenario 1:
                Questions: [First question about the topic] Also, [second related question about the same topic]
                Required Chunks: CHUNK_1, CHUNK_2, CHUNK_3, CHUNK_4

            """

            response = client.chat.completions.create(
                model="gpt-4.1",
                messages=[
                    {
                        "role": "system",
                        "content": "You are an expert at creating realistic customer service scenarios based on company policies.",
                    },
                    {"role": "user", "content": scenario_prompt},
                ],
                max_tokens=300,
                temperature=0.5,
            )

            scenario_text = response.choices[0].message.content.strip()

            # Parse scenarios from response
            scenarios = []
            scenario_chunks = {}  # Track which chunks each scenario needs
            lines = scenario_text.split("\n")
            current_scenario_num = 0

            for line in lines:
                line = line.strip()

                # Track which scenario we're currently parsing
                if line.startswith("Scenario 1:"):
                    current_scenario_num = 1

                # Look for lines that start with "Questions:" (multiple questions in one scenario)
                elif line.startswith("Questions:"):
                    questions = line.replace("Questions:", "").strip()
                    if questions:  # Only add non-empty questions
                        scenarios.append(questions)

                # Look for lines that start with "Required Chunks:"
                elif line.startswith("Required Chunks:") and current_scenario_num > 0:
                    chunks = line.replace("Required Chunks:", "").strip()
                    scenario_chunks[current_scenario_num] = chunks

            # Essential results logging
            print(f"✅ CREATED {len(scenarios)} SCENARIOS")
            for i, scenario in enumerate(scenarios, 1):
                chunks_needed = scenario_chunks.get(i, "Not specified")
                print(f"📝 Scenario {i}: {scenario}")
                print(f"📋 Required Chunks: {chunks_needed}")

            return scenarios, scenario_chunks

        except Exception as e:
            print(f"[ERROR] Error creating policy scenarios: {str(e)}")
            return [
                f"General {theme} inquiry for {department} with multiple questions about policies"
            ], {1: "ALL_CHUNKS"}


def main():
    # Function definitions first
    def format_display_name(name):
        """Convert camelCase or PascalCase to spaced words"""
        import re

        if not name:
            return name

        # Add space before uppercase letters that follow lowercase letters
        formatted = re.sub(r"([a-z])([A-Z])", r"\1 \2", name)

        # Handle specific cases - match the actual data
        formatted = formatted.replace("Cancelfilght", "Cancel Flight")  # Fixed typo
        formatted = formatted.replace(
            "Flightbooking", "Flight Booking"
        )  # Fixed capitalization
        formatted = formatted.replace("FlightBooking", "Flight Booking")
        formatted = formatted.replace("Rebooking", "Re-booking")

        return formatted

    if st.session_state.get("logged_in", False):
        # Logout functionality moved to bottom of main sidebar
        pass

    # --- DYNAMIC LOGIN SYSTEM ---
    if "logged_in" not in st.session_state:
        st.session_state.logged_in = False
        st.session_state.username = ""
        st.session_state.role = ""

    # Initialize evaluate session button state
    if "evaluate_session_clicked" not in st.session_state:
        st.session_state.evaluate_session_clicked = False
    if "evaluation_in_progress" not in st.session_state:
        st.session_state.evaluation_in_progress = False

    if not st.session_state.logged_in:
        st.title("🔐 Login to AI Customer Support Trainer")

        # Load dynamic users from GCS
        USERS = user_sim_manager.load_users()

        with st.form("login_form"):
            username = st.text_input("Username")
            password = st.text_input("Password", type="password")
            submitted = st.form_submit_button("Login")

        if submitted:
            user = USERS.get(username)
            if user and user["password"] == password:
                st.session_state.logged_in = True
                st.session_state.username = username
                st.session_state.role = user["role"]
                # Restore user data if exists (from session and disk)
                if "user_data" not in st.session_state:
                    st.session_state.user_data = {}
                user_data = st.session_state.user_data.get(username, {})
                # Try to load eval_summary and chat_history from disk if available
                user_dir = os.path.join("user_data", username)
                eval_summary = []
                chat_history = []
                if os.path.isdir(user_dir):
                    # Load all Excel files in user_dir except chat_history.xlsx
                    for fname in os.listdir(user_dir):
                        if fname.endswith(".xlsx") and fname != "chat_history.xlsx":
                            fpath = os.path.join(user_dir, fname)
                            with open(fpath, "rb") as f:
                                excel_bytes = f.read()
                            # Try to parse timestamp from filename
                            ts_match = re.search(r"eval_\d+_(.*?)\.xlsx", fname)
                            timestamp = (
                                ts_match.group(1).replace("_", " ").replace("-", ":")
                                if ts_match
                                else ""
                            )
                            # Read metadata from Excel file
                            try:
                                wb = openpyxl.load_workbook(fpath, data_only=True)
                                meta = wb["Metadata"]
                                meta_row = list(
                                    meta.iter_rows(
                                        min_row=2, max_row=2, values_only=True
                                    )
                                )[0]
                                (
                                    username_xl,
                                    dept_xl,
                                    issue_xl,
                                    tone_xl,
                                    ts_xl,
                                    score_xl,
                                ) = meta_row
                            except Exception:
                                (
                                    username_xl,
                                    dept_xl,
                                    issue_xl,
                                    tone_xl,
                                    ts_xl,
                                    score_xl,
                                ) = (username, "", "", "", timestamp, "")
                            eval_summary.append(
                                {
                                    "username": username_xl or username,
                                    "timestamp": ts_xl or timestamp,
                                    "excel_bytes": excel_bytes,
                                    "department": dept_xl or "",
                                    "issue_type": issue_xl or "",
                                    "tone": tone_xl or "",
                                    "eval_score": score_xl or "",
                                }
                            )
                    # Note: chat_history.xlsx loading removed - using session state only
                else:
                    eval_summary = user_data.get("eval_summary", [])
                    chat_history = user_data.get("chat_history", [])
                st.session_state.eval_summary = eval_summary
                # NEW: Only restore chat_history if not cleared by Start New Training Session
                if st.session_state.get("clear_chat_on_login", False):
                    st.session_state.chat_history = []
                    st.session_state["clear_chat_on_login"] = False
                else:
                    st.session_state.chat_history = chat_history
                st.success(f"Welcome, {username}!")
                st.rerun()
            else:
                st.error("Invalid username or password.")
        st.stop()

    # --- EVAL SUMMARY & CHAT HISTORY PERSISTENCE ---
    # Ensure eval_summary and chat_history exist in session
    if "eval_summary" not in st.session_state:
        st.session_state.eval_summary = []
    if "chat_history" not in st.session_state:
        st.session_state.chat_history = []

    # --- ROLE-BASED UI ---
    if st.session_state.role == "trainer":
        st.title("📊 Supervisor (Trainer) Dashboard")
        st.markdown("Monitor and analyze trainee performance across all sessions")

        # 🚀 PERFORMANCE OPTIMIZATION: In-memory caching implemented
        # - First load: Fetches from GCP bucket and caches in session state
        # - Subsequent loads: Uses cached data for instant response
        # - Manual refresh available to update cache with fresh data
        # - Cache automatically cleared on logout to free memory

        # Check for logout in sidebar FIRST (before any data loading)
        logout_requested = False
        refresh_requested = False
        with st.sidebar:
            st.header("🔍 Analytics Filters")

            # Create a placeholder for filters that will be populated after data loading
            filters_placeholder = st.empty()

            # Cache management buttons
            st.markdown("---")
            st.markdown("**📊 Data Management**")

            # Show cache status
            cached_count = len(st.session_state.get("supervisor_all_trainee_data", []))
            if cached_count > 0:
                st.success(f"✅ {cached_count} evaluations cached")
            else:
                st.info("🔄 No data cached")

            # Refresh data button
            if st.button(
                "🔄 Refresh Data",
                key="supervisor_refresh_btn",
                help="Clear cache and reload fresh data from cloud storage",
            ):
                refresh_requested = True

            # Logout button at bottom of sidebar
            st.markdown("---")
            if st.button(
                "🚪 Logout", key="supervisor_sidebar_logout_btn", type="secondary"
            ):
                logout_requested = True

        # Handle refresh BEFORE loading data
        if refresh_requested:
            # Clear the cached supervisor data to force fresh load
            if "supervisor_all_trainee_data" in st.session_state:
                del st.session_state["supervisor_all_trainee_data"]
            # Clear cached function data safely (Streamlit cache)
            try:
                from supervisor_dashboard import cached_users, cached_assignments

                cached_users.clear()
                cached_assignments.clear()
            except Exception:
                pass
            st.success("🔄 Cache cleared! Refreshing data...")
            st.rerun()

        # Handle logout BEFORE any data loading
        if logout_requested:
            # 🚀 COMPREHENSIVE SUPERVISOR LOGOUT CLEANUP:
            # - Login credentials
            # - Supervisor cached analytics data
            # - Policy-related session data
            # - Any trainee session data (in case supervisor used trainee features)
            print(
                f"[DEBUG] Supervisor {st.session_state.get('username', '')} logging out"
            )
            st.session_state.logged_in = False
            st.session_state.username = ""
            st.session_state.role = ""

            # 🚀 Clear supervisor cached data to free memory
            if "supervisor_all_trainee_data" in st.session_state:
                cached_count = len(st.session_state["supervisor_all_trainee_data"])
                del st.session_state["supervisor_all_trainee_data"]
            cached_supervisor_assignments.clear()

            # 🚀 Clear supervisor dashboard caches too
            try:
                from supervisor_dashboard import cached_users, cached_assignments

                cached_users.clear()
                cached_assignments.clear()
            except Exception:
                pass

            # 🚀 Clear policy-related session data (in case supervisor was also using trainee features)
            policy_keys = [
                "policy_scenarios",
                "current_policy_session",
                "policy_switch_scenario",
                "selected_policies",
                "scenario_chunks",
                "policy_context",
                "current_scenario",
                "messages",
                "gcs_evaluations",
                "evaluate_session_clicked",  # Clear evaluate session button state
            ]
            for key in policy_keys:
                if key in st.session_state:
                    if key == "policy_scenarios":
                        st.session_state[key] = []  # Reset to empty list
                    elif key == "current_policy_session":
                        st.session_state[key] = None  # Reset to None
                    else:
                        del st.session_state[key]  # Delete completely

            #  Clear GCS storage manager cache - FIXES THE ORIGINAL ISSUE!
            # This ensures fresh data loading on next login
            try:
                get_gcs_storage_manager.clear()
            except Exception as e:
                print(f"[DEBUG] ⚠️ Could not clear GCS cache: {e}")

            st.success("✅ Logged out successfully!")
            st.rerun()

        # Function to load all trainee data from GCS with in-memory caching
        def load_all_trainee_data_from_gcs():
            """🚀 Load all trainee evaluation data from GCS bucket with caching"""
            # Check if data is already cached in session state
            if "supervisor_all_trainee_data" in st.session_state:
                cached_data = st.session_state["supervisor_all_trainee_data"]
                return cached_data

            # If not cached, load from GCS
            gcs_manager = get_gcs_storage_manager()
            if not gcs_manager.is_connected():
                print("[ERROR] GCS not connected for trainer analytics")
                return []

            all_trainee_data = gcs_manager.load_all_evaluations_for_analytics()

            # Cache the data in session state for faster subsequent loads
            st.session_state["supervisor_all_trainee_data"] = all_trainee_data

            return all_trainee_data

        # 🚀 Load all trainee data from GCS with caching (ONLY if not logging out)
        with st.spinner("Loading analytics data from cloud storage..."):
            all_trainee_data = load_all_trainee_data_from_gcs()

        if not all_trainee_data:
            st.info(
                "📊 No evaluation data yet. Tabs are available - create simulations and complete training sessions to see analytics."
            )
            # Don't stop - let tabs show with empty filters
        else:
            # Show data source info
            unique_trainees = len(set([d["username"] for d in all_trainee_data]))
            st.success(
                f"📊 Loaded {len(all_trainee_data)} evaluations from {unique_trainees} trainee(s) from cloud storage"
            )

        # Now populate the sidebar filters with actual data
        with filters_placeholder.container():
            # Get unique values for filters
            all_usernames = sorted(list(set([d["username"] for d in all_trainee_data])))
            all_departments = sorted(
                list(
                    set([d["department"] for d in all_trainee_data if d["department"]])
                )
            )
            all_tones = sorted(
                list(set([d["tone"] for d in all_trainee_data if d["tone"]]))
            )

            # Filter controls
            selected_trainees = st.multiselect(
                "Select Trainees:", all_usernames, default=all_usernames
            )
            selected_departments = st.multiselect(
                "Select Departments:", all_departments, default=all_departments
            )
            selected_tones = st.multiselect(
                "Select Tones:", all_tones, default=all_tones
            )

            # Date range filter
            st.markdown("**Date Range:**")
            date_filter = st.selectbox(
                "Filter by:", ["All Time", "Last 7 Days", "Last 30 Days", "Today"]
            )

        # Apply filters
        filtered_data = all_trainee_data
        if selected_trainees:
            filtered_data = [
                d for d in filtered_data if d["username"] in selected_trainees
            ]
        if selected_departments:
            filtered_data = [
                d for d in filtered_data if d["department"] in selected_departments
            ]
        if selected_tones:
            filtered_data = [d for d in filtered_data if d["tone"] in selected_tones]

        # Main dashboard tabs
        tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs(
            [
                "👤 Associate Management",
                "📋 Training Session Assignment",
                "👥 Individual Performance",
                "🧠 Behaviour Analysis",
                "📊 Comparative Analysis",
                "📋 Detailed Reports",
            ]
        )

        with tab1:  # Associate Management Tab
            render_user_management_tab()

        with tab2:  # Training Session Assignment Tab
            render_simulation_assignment_tab()

        with tab3:  # Individual Performance Tab
            if filtered_data:
                # Trainee selector
                selected_trainee = st.selectbox(
                    "Select Trainee for Detailed View:",
                    sorted(list(set([d["username"] for d in filtered_data]))),
                    key="tab1_trainee_selector",
                )
                if selected_trainee:
                    trainee_data = [
                        d for d in filtered_data if d["username"] == selected_trainee
                    ]

                    # Trainee metrics
                    col1, col2, col3, col4 = st.columns(4)

                    with col1:
                        st.metric("Total Sessions", len(trainee_data))

                    with col2:
                        avg_score = np.mean(
                            [format_eval_score(d["eval_score"]) for d in trainee_data]
                        )
                        st.metric("Average Score", f"{avg_score:.1f}%")

                    with col3:
                        best_score = max(
                            [format_eval_score(d["eval_score"]) for d in trainee_data]
                        )
                        st.metric("Best Score", f"{best_score:.1f}%")

                    with col4:
                        latest_score = format_eval_score(
                            trainee_data[0]["eval_score"]
                        )  # Data is sorted by timestamp desc
                        st.metric("Latest Score", f"{latest_score:.1f}%")

                    # Show individual trainee progress
                    st.markdown("### 📈 Progress Over Time")
                    st.markdown("---")
                    st.markdown("#### � Session Overview")
                    trainee_df = pd.DataFrame(trainee_data)
                    trainee_df["timestamp"] = pd.to_datetime(trainee_df["timestamp"])
                    trainee_df = trainee_df.sort_values("timestamp")
                    trainee_df["session_number"] = range(1, len(trainee_df) + 1)
                    # Format eval_score to numeric values
                    trainee_df["eval_score"] = trainee_df["eval_score"].apply(
                        format_eval_score
                    )

                    fig, ax = plt.subplots(figsize=(12, 6))

                    sns.lineplot(
                        data=trainee_df,
                        x="session_number",
                        y="eval_score",
                        marker="o",
                        markersize=8,
                        linewidth=3,
                        color="#2E86AB",
                        ax=ax,
                    )

                    # Fill area under curve
                    ax.fill_between(
                        trainee_df["session_number"],
                        trainee_df["eval_score"],
                        alpha=0.3,
                        color="#2E86AB",
                    )

                    # Add value annotations
                    for x, y in zip(
                        trainee_df["session_number"], trainee_df["eval_score"]
                    ):
                        ax.annotate(
                            f"{y:.1f}%",
                            (x, y),
                            textcoords="offset points",
                            xytext=(0, 10),
                            ha="center",
                            fontweight="bold",
                            bbox=dict(
                                boxstyle="round,pad=0.3", facecolor="white", alpha=0.8
                            ),
                        )

                    ax.set_xlabel("Session Number", fontsize=12, fontweight="bold")
                    ax.set_ylabel(
                        "Evaluation Score (%)", fontsize=12, fontweight="bold"
                    )
                    ax.set_title(
                        f"{selected_trainee}'s Performance Progress",
                        fontsize=14,
                        fontweight="bold",
                        pad=20,
                    )
                    ax.set_ylim(0, 105)
                    ax.grid(True, linestyle="--", alpha=0.7)

                    # Add average line
                    ax.axhline(
                        y=avg_score, color="red", linestyle="--", alpha=0.7, linewidth=2
                    )
                    ax.text(
                        len(trainee_df),
                        avg_score,
                        f"Avg: {avg_score:.1f}%",
                        verticalalignment="bottom",
                        color="red",
                        fontweight="bold",
                    )

                    plt.tight_layout()
                    st.pyplot(fig)
                    plt.close()

                    # Add Behaviors Graph for Individual Trainee
                    st.markdown("---")
                    st.markdown("### 🏆 Individual Behavior Performance")

                    # Create behaviors graph for this specific trainee
                    create_behaviors_graph(trainee_data, f"{selected_trainee}'s ")

        with tab4:  # Behaviour Analysis Tab
            st.markdown("### 🧠 Behavior Analysis")
            st.markdown("Compare all associates' performance on a specific behavior")

            if filtered_data:

                # Extract all available behaviors
                all_behaviors = set()
                for record in filtered_data:
                    excel_bytes = record.get("excel_bytes")
                    if excel_bytes:
                        try:
                            with BytesIO(excel_bytes) as io:
                                wb = openpyxl.load_workbook(io, data_only=True)
                                if "Evaluation Table" in wb.sheetnames:
                                    ws = wb["Evaluation Table"]
                                    for row in ws.iter_rows(
                                        min_row=2, values_only=True
                                    ):
                                        if len(row) >= 4:
                                            behavior = row[0]
                                            if behavior and behavior.strip():
                                                all_behaviors.add(behavior.strip())
                        except Exception:
                            continue

                if all_behaviors:
                    behavior_list = sorted(list(all_behaviors))
                    selected_behavior = st.selectbox(
                        "Select Behavior to Compare Across Associates:",
                        [""] + behavior_list,
                        key="behavior_comparison_select",
                    )

                    if selected_behavior:
                        # Analyze this behavior across all associates
                        associate_behavior_data = {}

                        for record in filtered_data:
                            username = record.get("username", "Unknown")
                            excel_bytes = record.get("excel_bytes")

                            if excel_bytes and username != "Unknown":
                                if username not in associate_behavior_data:
                                    associate_behavior_data[username] = {
                                        "scores": [],
                                        "max_scores": [],
                                        "sessions": 0,
                                    }

                                try:
                                    with BytesIO(excel_bytes) as io:
                                        wb = openpyxl.load_workbook(io, data_only=True)
                                        if "Evaluation Table" in wb.sheetnames:
                                            ws = wb["Evaluation Table"]
                                            for row in ws.iter_rows(
                                                min_row=2, values_only=True
                                            ):
                                                if len(row) >= 4:
                                                    behavior, _, max_score, score = row[
                                                        :4
                                                    ]
                                                    if (
                                                        behavior
                                                        and behavior.strip()
                                                        == selected_behavior
                                                    ):
                                                        try:
                                                            score = float(score)
                                                            max_score = float(max_score)
                                                            associate_behavior_data[
                                                                username
                                                            ]["scores"].append(score)
                                                            associate_behavior_data[
                                                                username
                                                            ]["max_scores"].append(
                                                                max_score
                                                            )
                                                            associate_behavior_data[
                                                                username
                                                            ]["sessions"] += 1
                                                        except (ValueError, TypeError):
                                                            continue
                                except Exception:
                                    continue

                        # Process the data for comparison
                        comparison_data = []
                        for associate, data in associate_behavior_data.items():
                            if data["scores"]:  # Has data for this behavior
                                avg_score = sum(data["scores"]) / len(data["scores"])
                                avg_max = sum(data["max_scores"]) / len(
                                    data["max_scores"]
                                )
                                percentage = (
                                    (avg_score / avg_max * 100) if avg_max > 0 else 0
                                )
                                comparison_data.append(
                                    {
                                        "Associate": associate,
                                        "Avg_Score": avg_score,
                                        "Avg_Max": avg_max,
                                        "Percentage": percentage,
                                        "Sessions": data["sessions"],
                                        "Has_Data": True,
                                    }
                                )
                            else:  # No data for this behavior
                                comparison_data.append(
                                    {
                                        "Associate": associate,
                                        "Avg_Score": 0,
                                        "Avg_Max": 0,
                                        "Percentage": 0,
                                        "Sessions": 0,
                                        "Has_Data": False,
                                    }
                                )

                        # Only include associates who have data for this behavior
                        # (Removed the section that adds associates without data)

                        if comparison_data:
                            st.markdown(
                                f"### 📊 Comparison Results for: **{selected_behavior}**"
                            )

                            # Summary statistics - only show associates with data
                            associates_with_data = [
                                d for d in comparison_data if d["Has_Data"]
                            ]

                            col1, col2 = st.columns(2)
                            with col1:
                                st.metric(
                                    "Associates with Data", len(associates_with_data)
                                )
                            with col2:
                                if associates_with_data:
                                    avg_performance = sum(
                                        d["Percentage"] for d in associates_with_data
                                    ) / len(associates_with_data)
                                    st.metric(
                                        "Average Performance", f"{avg_performance:.1f}%"
                                    )

                            # Create comparison visualization
                            if associates_with_data:
                                # Sort by percentage for better visualization
                                associates_with_data.sort(
                                    key=lambda x: x["Percentage"], reverse=True
                                )

                                # Add filter dropdown if more than 5 associates
                                display_data = associates_with_data
                                if len(associates_with_data) > 5:
                                    filter_option = st.selectbox(
                                        "Display Option:",
                                        [
                                            "All Associates",
                                            "Top 5 Performers",
                                            "Bottom 5 Performers",
                                        ],
                                        key="behavior_display_filter",
                                    )

                                    if filter_option == "Top 5 Performers":
                                        display_data = associates_with_data[:5]
                                    elif filter_option == "Bottom 5 Performers":
                                        display_data = associates_with_data[-5:]

                                # Create DataFrame for plotting - only selected associates
                                plot_data = []
                                for item in display_data:
                                    plot_data.append(
                                        {
                                            "Associate": item["Associate"],
                                            "Percentage": item["Percentage"],
                                            "Sessions": item["Sessions"],
                                        }
                                    )

                                df_plot = pd.DataFrame(plot_data)

                                # Check if all values are zero (for bottom performers)
                                all_zero = all(
                                    item["Percentage"] == 0 for item in display_data
                                )

                                if all_zero:
                                    st.warning(
                                        f"⚠️ All selected associates have 0% performance for '{selected_behavior}'. No graph to display."
                                    )

                                    # Show a simple info table instead
                                    st.markdown("**Associates with 0% Performance:**")
                                    zero_df = pd.DataFrame(
                                        [
                                            {
                                                "Associate": item["Associate"],
                                                "Score": f"{item['Avg_Score']:.1f}/{item['Avg_Max']:.1f}",
                                                "Sessions": item["Sessions"],
                                            }
                                            for item in display_data
                                        ]
                                    )
                                    st.dataframe(zero_df, use_container_width=True)
                                else:
                                    # Create vertical bar plot with improved styling
                                    plt.style.use("default")  # Reset to clean style

                                    # Simple figure sizing based on number of associates
                                    num_associates = len(display_data)
                                    if num_associates == 1:
                                        # Bigger size for single bar with more breathing room
                                        fig, ax = plt.subplots(
                                            figsize=(8, 8), facecolor="white"
                                        )
                                    else:
                                        # Dynamic sizing for multiple bars
                                        fig_width = max(10, num_associates * 1.5)
                                        fig, ax = plt.subplots(
                                            figsize=(fig_width, 8), facecolor="white"
                                        )

                                    # Use improved color scheme with better contrast
                                    colors = []
                                    edge_colors = []
                                    for percentage in df_plot["Percentage"]:
                                        if percentage >= 80:
                                            colors.append(
                                                "#2E7D32"
                                            )  # Darker green for better visibility
                                            edge_colors.append("#1B5E20")
                                        elif percentage >= 60:
                                            colors.append(
                                                "#F57C00"
                                            )  # Darker orange for better visibility
                                            edge_colors.append("#E65100")
                                        else:
                                            colors.append(
                                                "#C62828"
                                            )  # Darker red for better visibility
                                            edge_colors.append("#B71C1C")

                                    # Simple bar width logic
                                    if num_associates == 1:
                                        bar_width = (
                                            0.3  # Simple fixed width for single bar
                                        )
                                    else:
                                        bar_width = (
                                            0.8  # Standard width for multiple bars
                                        )

                                    bars = ax.bar(
                                        range(len(df_plot)),
                                        df_plot["Percentage"],
                                        width=bar_width,
                                        color=colors,
                                        edgecolor=edge_colors,
                                        linewidth=1.5,
                                        alpha=0.9,
                                    )

                                    # Customize axes
                                    ax.set_xticks(range(len(df_plot)))
                                    ax.set_xticklabels(
                                        df_plot["Associate"],
                                        rotation=45,
                                        ha="right",
                                        fontsize=11,
                                        fontweight="bold",
                                    )
                                    ax.set_ylabel(
                                        "Performance Percentage (%)",
                                        fontsize=12,
                                        fontweight="bold",
                                    )
                                    ax.set_title(
                                        f"Associate Comparison: {selected_behavior}",
                                        fontsize=14,
                                        fontweight="bold",
                                        pad=20,
                                    )

                                    # Add percentage labels on top of bars with better styling
                                    for i, (bar, item) in enumerate(
                                        zip(bars, display_data)
                                    ):
                                        height = bar.get_height()
                                        if (
                                            height > 0
                                        ):  # Only add labels for non-zero bars
                                            ax.text(
                                                bar.get_x() + bar.get_width() / 2.0,
                                                height
                                                + (
                                                    max(df_plot["Percentage"]) * 0.02
                                                ),  # Dynamic spacing
                                                f"{item['Percentage']:.1f}%\n({item['Sessions']} sessions)",
                                                ha="center",
                                                va="bottom",
                                                fontsize=10,
                                                fontweight="bold",
                                                bbox=dict(
                                                    boxstyle="round,pad=0.3",
                                                    facecolor="white",
                                                    alpha=0.8,
                                                ),
                                            )

                                    # Improve grid styling
                                    ax.grid(
                                        True,
                                        linestyle="-",
                                        alpha=0.3,
                                        axis="y",
                                        color="gray",
                                    )
                                    ax.set_axisbelow(True)

                                    # Set y-axis limit to give space for labels
                                    max_percentage = max(df_plot["Percentage"])
                                    ax.set_ylim(0, max_percentage * 1.2)

                                    # Add performance threshold lines with better visibility
                                    # Always show threshold lines for context
                                    ax.axhline(
                                        y=80,
                                        color="#2E7D32",
                                        linestyle="--",
                                        alpha=0.8,
                                        linewidth=2,
                                        label="Excellent (80%+)",
                                    )
                                    ax.axhline(
                                        y=60,
                                        color="#F57C00",
                                        linestyle="--",
                                        alpha=0.8,
                                        linewidth=2,
                                        label="Satisfactory (60%+)",
                                    )

                                    # Add a line for poor performance threshold
                                    ax.axhline(
                                        y=0,
                                        color="#C62828",
                                        linestyle="--",
                                        alpha=0.8,
                                        linewidth=2,
                                        label="Poor Performance (<60%)",
                                    )

                                    # Improved legend with better positioning based on number of bars
                                    if num_associates == 1:
                                        # For single bar, place legend in upper right with more space
                                        legend = ax.legend(
                                            loc="upper right",
                                            fontsize=11,
                                            frameon=True,
                                            fancybox=True,
                                            shadow=True,
                                            framealpha=0.95,
                                            facecolor="white",
                                            edgecolor="black",
                                            bbox_to_anchor=(
                                                0.98,
                                                0.98,
                                            ),  # Slight padding from edges
                                        )
                                    else:
                                        # For multiple bars, use upper right inside plot
                                        legend = ax.legend(
                                            loc="upper right",
                                            fontsize=12,
                                            frameon=True,
                                            fancybox=True,
                                            shadow=True,
                                            framealpha=0.95,
                                            facecolor="white",
                                            edgecolor="black",
                                        )
                                    legend.get_frame().set_linewidth(1.5)

                                    # Style the plot area
                                    ax.spines["top"].set_visible(False)
                                    ax.spines["right"].set_visible(False)
                                    ax.spines["left"].set_linewidth(1.5)
                                    ax.spines["bottom"].set_linewidth(1.5)

                                    # Improve tick styling
                                    ax.tick_params(
                                        axis="y", labelsize=11, colors="black"
                                    )
                                    ax.tick_params(
                                        axis="x", labelsize=11, colors="black"
                                    )

                                    plt.tight_layout()
                                    st.pyplot(fig)
                                    plt.close()
                            else:
                                st.warning(
                                    f"⚠️ No associates have evaluation data for '{selected_behavior}'"
                                )
                        else:
                            st.info("ℹ️ No comparison data available")
                else:
                    st.warning("⚠️ No behaviors found in evaluation data")
            else:
                st.warning(
                    "⚠️ No evaluation data available. Associates need to complete evaluations first."
                )

        with tab5:  # Comparative Analysis Tab
            st.markdown("### 🆚 Comparative Analysis")

            if filtered_data:
                # Trainee comparison
                st.markdown("### 👥 Trainee Performance Comparison")

                trainee_stats = {}
                for d in filtered_data:
                    username = d["username"]
                    if username not in trainee_stats:
                        trainee_stats[username] = []
                    trainee_stats[username].append(format_eval_score(d["eval_score"]))

                comparison_data = []
                for username, scores in trainee_stats.items():
                    comparison_data.append(
                        {
                            "Trainee": username,
                            "Sessions": len(scores),
                            "Average Score": np.mean(scores),
                            "Best Score": max(scores),
                            "Latest Score": scores[0],  # Assuming first is latest
                            "Improvement": (
                                scores[0] - scores[-1] if len(scores) > 1 else 0
                            ),
                        }
                    )

                comparison_df = pd.DataFrame(comparison_data)
                comparison_df = comparison_df.sort_values(
                    "Average Score", ascending=False
                )

                # Display comparison table
                st.dataframe(comparison_df, use_container_width=True)

                # Comparison chart
                fig, ax = plt.subplots(figsize=(12, 8))

                x_pos = np.arange(len(comparison_df))
                bars = ax.bar(
                    x_pos, comparison_df["Average Score"], color="skyblue", alpha=0.7
                )

                # Add value labels on bars
                for i, (bar, score) in enumerate(
                    zip(bars, comparison_df["Average Score"])
                ):
                    ax.text(
                        bar.get_x() + bar.get_width() / 2,
                        bar.get_height() + 1,
                        f"{score:.1f}%",
                        ha="center",
                        va="bottom",
                        fontweight="bold",
                    )

                ax.set_xlabel("Trainees", fontsize=12, fontweight="bold")
                ax.set_ylabel("Average Score (%)", fontsize=12, fontweight="bold")
                ax.set_title(
                    "Trainee Performance Comparison",
                    fontsize=14,
                    fontweight="bold",
                    pad=20,
                )
                ax.set_xticks(x_pos)
                ax.set_xticklabels(comparison_df["Trainee"], rotation=45, ha="right")
                ax.set_ylim(0, 105)
                ax.grid(True, axis="y", linestyle="--", alpha=0.7)

                plt.tight_layout()
                st.pyplot(fig)
                plt.close()

                st.markdown("---")
                st.markdown("### 🏢 Department Performance")

                dept_scores = defaultdict(list)
                for d in filtered_data:
                    if d["department"]:
                        dept_scores[d["department"]].append(
                            format_eval_score(d["eval_score"])
                        )

                if dept_scores:
                    dept_avg = {
                        dept: np.mean(scores) for dept, scores in dept_scores.items()
                    }
                    dept_df = pd.DataFrame(
                        list(dept_avg.items()), columns=["Department", "Avg Score"]
                    )
                    dept_df = dept_df.sort_values("Avg Score", ascending=True)

                    fig, ax = plt.subplots(figsize=(12, 6))
                    sns.barplot(
                        data=dept_df,
                        y="Department",
                        x="Avg Score",
                        palette="viridis",
                        ax=ax,
                    )

                    # Add value labels
                    for i, v in enumerate(dept_df["Avg Score"]):
                        ax.text(v + 1, i, f"{v:.1f}%", va="center", fontweight="bold")

                    ax.set_xlabel("Average Score (%)", fontsize=12, fontweight="bold")
                    ax.set_ylabel("Department", fontsize=12, fontweight="bold")
                    ax.set_title(
                        "Average Performance by Department",
                        fontsize=14,
                        fontweight="bold",
                    )
                    ax.set_xlim(0, 105)
                    plt.tight_layout()
                    st.pyplot(fig)
                    plt.close()

        with tab6:  # Detailed Reports Tab
            # Check if we're in supervisor report view mode
            if "view_supervisor_eval_idx" in st.session_state:
                st.markdown("### 📊 Detailed Evaluation Report")

                # Back button
                if st.button(
                    "← Back to Overall Reports", key="back_to_supervisor_reports"
                ):
                    st.session_state.pop("view_supervisor_eval_idx", None)
                    st.session_state.pop("supervisor_eval_data", None)
                    st.rerun()

                eval_data = st.session_state.get("supervisor_eval_data")
                if eval_data:
                    # Enhanced CSS for the detailed view (same as trainee view)
                    st.markdown(
                        """
                        <style>
                        .detail-container {
                            background: linear-gradient(135deg, #f5f7fa 0%, #c3cfe2 100%);
                            border-radius: 15px;
                            padding: 2rem;
                            margin: 1rem 0;
                            box-shadow: 0 8px 32px rgba(0, 0, 0, 0.1);
                            border: 1px solid rgba(255, 255, 255, 0.2);
                        }
                        .detail-header {
                            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                            color: white;
                            padding: 1.5rem;
                            border-radius: 12px;
                            text-align: center;
                            margin-bottom: 2rem;
                            box-shadow: 0 4px 15px rgba(102, 126, 234, 0.3);
                        }
                        .sheet-container {
                            background: white;
                            padding: 1rem;
                            margin: 0.5rem 0;
                            border-left: 3px solid #667eea;
                        }
                        .sheet-title {
                            color: #2d3748;
                            font-size: 1.3rem;
                            font-weight: 600;
                            margin-bottom: 1rem;
                            padding-bottom: 0.5rem;
                            border-bottom: 2px solid #e2e8f0;
                            display: flex;
                            align-items: center;
                            gap: 0.5rem;
                        }
                        .chat-message {
                            background: #f8fafc;
                            border-radius: 6px;
                            padding: 0.6rem 0.8rem;
                            margin: 0.3rem 0;
                            border-left: 3px solid #4299e1;
                            font-size: 0.9rem;
                            line-height: 1.4;
                        }
                        .chat-message.user {
                            background: #e6f3ff;
                            border-left-color: #3182ce;
                        }
                        .chat-message.assistant {
                            background: #f0f9ff;
                            border-left-color: #0ea5e9;
                        }
                        .chat-role {
                            font-weight: 600;
                            font-size: 0.75rem;
                            color: #4a5568;
                            margin-bottom: 0.3rem;
                            text-transform: uppercase;
                            letter-spacing: 0.3px;
                        }
                        .behavior-score {
                            display: inline-block;
                            background: linear-gradient(135deg, #48bb78 0%, #38a169 100%);
                            color: white;
                            padding: 0.3rem 1rem;
                            border-radius: 20px;
                            font-weight: 600;
                            font-size: 0.9rem;
                            margin: 0.2rem;
                            box-shadow: 0 2px 4px rgba(72, 187, 120, 0.3);
                        }
                        .behavior-item {
                            background: #f9fafb;
                            border-radius: 4px;
                            padding: 0.4rem;
                            margin: 0.2rem 0;
                            border-left: 3px solid #e5e7eb;
                            font-size: 0.85rem;
                        }
                        .behavior-item:hover {
                            background: #f3f4f6;
                            border-color: #d1d5db;
                        }
                        .metadata-card {
                            background: linear-gradient(135deg, #fef5e7 0%, #fed7aa 100%);
                            border-radius: 10px;
                            padding: 1rem;
                            margin: 1rem 0;
                            border-left: 4px solid #f59e0b;
                        }
                        </style>
                        """,
                        unsafe_allow_html=True,
                    )

                    st.markdown(
                        '<div class="detail-container">', unsafe_allow_html=True
                    )
                    st.markdown(
                        f'<div class="detail-header"><h2>📊 Detailed Evaluation Report - {eval_data.get("username", "N/A")}</h2></div>',
                        unsafe_allow_html=True,
                    )

                    # Display session metadata in a card
                    st.markdown(
                        f"""
                        <div class="metadata-card">
                            <strong>📝 Session Information:</strong><br>
                            <strong>Trainee:</strong> {eval_data.get('username', 'N/A')} |
                            <strong>Department:</strong> {eval_data.get('department', 'N/A')} |
                            <strong>Issue Type:</strong> {eval_data.get('issue_type', 'N/A')}<br>
                            <strong>Tone:</strong> {eval_data.get('tone', 'N/A')} |
                            <strong>Date:</strong> {eval_data.get('timestamp', 'N/A')} |
                            <strong>Score:</strong> <span class="behavior-score">{format_eval_score(eval_data.get('eval_score', 'N/A')):.1f}%</span>
                        </div>
                        """,
                        unsafe_allow_html=True,
                    )

                    excel_bytes = eval_data["excel_bytes"]
                    with BytesIO(excel_bytes) as excel_io:
                        wb = openpyxl.load_workbook(excel_io)

                        for sheet_name in wb.sheetnames:
                            st.markdown(
                                '<div class="sheet-container">', unsafe_allow_html=True
                            )

                            # Determine sheet icon
                            if (
                                "chat" in sheet_name.lower()
                                or "conversation" in sheet_name.lower()
                            ):
                                icon = "💬"
                            elif (
                                "evaluation" in sheet_name.lower()
                                or "behavior" in sheet_name.lower()
                            ):
                                icon = "📊"
                            elif "metadata" in sheet_name.lower():
                                icon = "📋"
                            else:
                                icon = "📄"

                            st.markdown(
                                f'<div class="sheet-title">{icon} {sheet_name}</div>',
                                unsafe_allow_html=True,
                            )

                            ws = wb[sheet_name]
                            data = list(ws.values)

                            if data:
                                df = pd.DataFrame(data[1:], columns=data[0])

                                # Special formatting for different sheet types
                                if (
                                    "chat" in sheet_name.lower()
                                    or "conversation" in sheet_name.lower()
                                ):
                                    # Enhanced chat history display
                                    st.markdown("📝 **Chat Conversation:**")

                                    if df.empty:
                                        st.info("No chat history found in this sheet.")
                                    else:
                                        for _, row in df.iterrows():
                                            # Try different possible column names
                                            role = ""
                                            content = ""

                                            # Common column names for role
                                            for role_col in [
                                                "Role",
                                                "role",
                                                "Speaker",
                                                "speaker",
                                                "User",
                                                "user",
                                            ]:
                                                if role_col in df.columns and pd.notna(
                                                    row.get(role_col)
                                                ):
                                                    role = str(row[role_col]).strip()
                                                    break

                                            # Common column names for content
                                            for content_col in [
                                                "Message",
                                                "message",
                                                "Content",
                                                "content",
                                                "Text",
                                                "text",
                                            ]:
                                                if (
                                                    content_col in df.columns
                                                    and pd.notna(row.get(content_col))
                                                ):
                                                    content = str(
                                                        row[content_col]
                                                    ).strip()
                                                    break

                                            if (
                                                content
                                            ):  # Only display if we have content
                                                role_display = (
                                                    role if role else "Unknown"
                                                )
                                                message_class = (
                                                    "user"
                                                    if "user" in role.lower()
                                                    or "customer" in role.lower()
                                                    else "assistant"
                                                )

                                                st.markdown(
                                                    f"""
                                                    <div class="chat-message {message_class}">
                                                        <div class="chat-role">{role_display}</div>
                                                        {content}
                                                    </div>
                                                    """,
                                                    unsafe_allow_html=True,
                                                )

                                elif (
                                    "evaluation" in sheet_name.lower()
                                    or "behavior" in sheet_name.lower()
                                    or "table" in sheet_name.lower()
                                ):
                                    # Enhanced evaluation table display - exact same as associates
                                    st.markdown("📊 **Behavior Evaluation Scores:**")

                                    if df.empty:
                                        st.info(
                                            "No evaluation data found in this sheet."
                                        )
                                    else:
                                        # Add column headers for 4-column layout
                                        st.markdown(
                                            """
                                            <div class="behavior-item" style="background: #667eea; color: white; margin-bottom: 0.5rem;">
                                                <table style="width: 100%; border-collapse: collapse; font-size: 0.85rem; table-layout: fixed;">
                                                    <tr>
                                                        <th style="font-weight: 600; padding: 0.6rem 0.5rem; text-align: center; width: 20%; border-right: 1px solid rgba(255,255,255,0.3);">
                                                            📋 Behavior
                                                        </th>
                                                        <th style="font-weight: 600; padding: 0.6rem 0.5rem; text-align: center; width: 15%; border-right: 1px solid rgba(255,255,255,0.3);">
                                                            🎯 Score
                                                        </th>
                                                        <th style="font-weight: 600; padding: 0.6rem 0.5rem; text-align: center; width: 30%; border-right: 1px solid rgba(255,255,255,0.3);">
                                                            💬 Citation
                                                        </th>
                                                        <th style="font-weight: 600; padding: 0.6rem 0.5rem; text-align: center; width: 35%;">
                                                            📝 Feedback
                                                        </th>
                                                    </tr>
                                                </table>
                                            </div>
                                            """,
                                            unsafe_allow_html=True,
                                        )

                                        for _, row in df.iterrows():
                                            # Try different possible column names with more flexibility
                                            behavior = ""
                                            criteria = ""
                                            max_score = ""
                                            score = ""
                                            feedback = ""
                                            reasoning = ""
                                            citation = ""

                                            # More comprehensive column matching
                                            for col in df.columns:
                                                col_str = str(col).strip()
                                                col_lower = col_str.lower()
                                                row_value = (
                                                    str(row[col]).strip()
                                                    if not pd.isna(row[col])
                                                    else ""
                                                )

                                                # Check for behavior column
                                                if not behavior and (
                                                    "behavior" in col_lower
                                                    or "behaviour" in col_lower
                                                ):
                                                    behavior = row_value

                                                # Check for criteria column
                                                elif not criteria and (
                                                    "criteria" in col_lower
                                                    or "criterion" in col_lower
                                                ):
                                                    criteria = row_value

                                                # Check for max score column
                                                elif not max_score and (
                                                    "max" in col_lower
                                                    and "score" in col_lower
                                                ):
                                                    max_score = row_value

                                                # Check for score column
                                                elif not score and (
                                                    col_lower == "score"
                                                    or (
                                                        col_lower.endswith("score")
                                                        and "max" not in col_lower
                                                    )
                                                ):
                                                    score = row_value

                                                # Check for feedback/reasoning column - prioritize "Reason" column
                                                elif not feedback and (
                                                    col_lower == "reason"
                                                    or col_lower == "reasons"
                                                    or any(
                                                        pattern in col_lower
                                                        for pattern in [
                                                            "feedback",
                                                            "comment",
                                                            "remarks",
                                                            "notes",
                                                            "suggestion",
                                                            "improvement",
                                                            "reasoning",
                                                        ]
                                                    )
                                                ):
                                                    feedback = row_value

                                                # Check for citation column
                                                elif not citation and any(
                                                    pattern in col_lower
                                                    for pattern in [
                                                        "citation",
                                                        "cite",
                                                        "reference",
                                                        "quote",
                                                        "example",
                                                    ]
                                                ):
                                                    citation = row_value

                                            # Fallback logic
                                            if not behavior and len(df.columns) > 0:
                                                behavior = (
                                                    str(row[df.columns[0]]).strip()
                                                    if not pd.isna(row[df.columns[0]])
                                                    else ""
                                                )

                                            # Try to find score in any numeric column
                                            if not score:
                                                for col in df.columns:
                                                    try:
                                                        val = str(row[col]).strip()
                                                        if (
                                                            val
                                                            and val != "nan"
                                                            and val.replace(".", "")
                                                            .replace(",", "")
                                                            .isdigit()
                                                        ):
                                                            score = val
                                                            break
                                                    except:
                                                        continue

                                            # Try to find feedback/reasoning in longer text columns - prioritize "Reason" column
                                            if not feedback or feedback in [
                                                "nan",
                                                "None",
                                                "",
                                                "No feedback provided",
                                                "Performance evaluation criteria",
                                            ]:
                                                # First, specifically look for "Reason" column
                                                for col in df.columns:
                                                    col_lower = str(col).lower()
                                                    if (
                                                        col_lower == "reason"
                                                        or col_lower == "reasons"
                                                    ):
                                                        val = (
                                                            str(row[col]).strip()
                                                            if not pd.isna(row[col])
                                                            else ""
                                                        )
                                                        if val and val != "nan":
                                                            feedback = val
                                                            break

                                                # If still no feedback, look for other text columns
                                                if not feedback or feedback in [
                                                    "nan",
                                                    "None",
                                                    "",
                                                ]:
                                                    for col in df.columns:
                                                        col_lower = str(col).lower()
                                                        if any(
                                                            skip in col_lower
                                                            for skip in [
                                                                "behavior",
                                                                "behaviour",
                                                                "criteria",
                                                                "score",
                                                                "max",
                                                                "citation",
                                                            ]
                                                        ):
                                                            continue

                                                        val = (
                                                            str(row[col]).strip()
                                                            if not pd.isna(row[col])
                                                            else ""
                                                        )
                                                        if (
                                                            val
                                                            and val != "nan"
                                                            and len(val) > 15
                                                        ):  # Look for substantial text
                                                            feedback = val
                                                            break

                                            # Skip completely empty rows
                                            if not behavior or behavior in [
                                                "nan",
                                                "None",
                                                "",
                                            ]:
                                                continue

                                            # Clean up values
                                            max_score = (
                                                max_score
                                                if max_score
                                                and max_score not in ["nan", "None", ""]
                                                else "5"
                                            )
                                            score = (
                                                score
                                                if score
                                                and score not in ["nan", "None", ""]
                                                else "0"
                                            )
                                            feedback = (
                                                feedback
                                                if feedback
                                                and feedback
                                                not in [
                                                    "nan",
                                                    "None",
                                                    "",
                                                    "Performance evaluation criteria",
                                                ]
                                                else ""
                                            )
                                            citation = (
                                                citation
                                                if citation
                                                and citation not in ["nan", "None", ""]
                                                else ""
                                            )

                                            # Create 4-column table layout with full text wrapping (no truncation)
                                            # Function to clean and prepare text for display
                                            def clean_text(text):
                                                if not text or text in [
                                                    "nan",
                                                    "None",
                                                    "",
                                                ]:
                                                    return "<em>No data provided</em>"
                                                return (
                                                    str(text)
                                                    .strip()
                                                    .replace("\n", "<br>")
                                                )

                                            clean_behavior = clean_text(behavior)
                                            clean_citation = clean_text(citation)
                                            clean_feedback = clean_text(feedback)

                                            st.markdown(
                                                f"""
                                                <div class="behavior-item">
                                                    <table style="width: 100%; border-collapse: collapse; font-size: 0.85rem; table-layout: fixed;">
                                                        <tr style="vertical-align: top;">
                                                            <td style="font-weight: 600; color: #2d3748; padding: 0.6rem 0.5rem; border-right: 1px solid #e5e7eb; width: 20%; word-wrap: break-word; overflow-wrap: break-word; white-space: normal; line-height: 1.4;">
                                                                🎯 {clean_behavior}
                                                            </td>
                                                            <td style="text-align: center; padding: 0.6rem 0.5rem; border-right: 1px solid #e5e7eb; width: 15%; word-wrap: break-word; vertical-align: middle;">
                                                                <span class="behavior-score">{score}/{max_score}</span>
                                                            </td>
                                                            <td style="padding: 0.6rem 0.5rem; border-right: 1px solid #e5e7eb; width: 30%; word-wrap: break-word; overflow-wrap: break-word; white-space: normal; font-style: italic; color: #4a5568; line-height: 1.4;">
                                                                💬 {clean_citation}
                                                            </td>
                                                            <td style="padding: 0.6rem 0.5rem; width: 35%; word-wrap: break-word; overflow-wrap: break-word; white-space: normal; line-height: 1.4;">
                                                                {clean_feedback}
                                                            </td>
                                                        </tr>
                                                    </table>
                                                </div>
                                                """,
                                                unsafe_allow_html=True,
                                            )

                                else:
                                    # Display other sheets as regular dataframes
                                    if df.empty:
                                        st.info(f"No data found in {sheet_name}.")
                                    else:
                                        st.dataframe(df, use_container_width=True)

                            st.markdown("</div>", unsafe_allow_html=True)

                    st.markdown("</div>", unsafe_allow_html=True)

                    # Add back button at the bottom
                    st.markdown("---")
                    if st.button(
                        "← Back to Overall Reports",
                        key="bottom_back_to_supervisor_reports",
                        type="primary",
                    ):
                        st.session_state.pop("view_supervisor_eval_idx", None)
                        st.session_state.pop("supervisor_eval_data", None)
                        st.rerun()
            else:
                # Regular detailed reports view
                st.markdown("### 📋 Detailed Evaluation Reports")

                if filtered_data:
                    # Search and filter options
                    col1, col2 = st.columns([2, 1])

                    with col1:
                        search_term = st.text_input(
                            "Search evaluations:",
                            placeholder="Enter trainee name, department, etc.",
                        )

                    with col2:
                        sort_by = st.selectbox(
                            "Sort by:",
                            ["Timestamp (Latest)", "Score (Highest)", "Trainee Name"],
                        )

                    # Filter and sort data
                    display_data = filtered_data

                    if search_term:
                        display_data = [
                            d
                            for d in display_data
                            if search_term.lower() in d["username"].lower()
                            or search_term.lower() in d["department"].lower()
                            or search_term.lower() in d["issue_type"].lower()
                        ]

                    # Sort data
                    if sort_by == "Score (Highest)":
                        display_data = sorted(
                            display_data, key=lambda x: x["eval_score"], reverse=True
                        )
                    elif sort_by == "Trainee Name":
                        display_data = sorted(display_data, key=lambda x: x["username"])
                    # Default is already timestamp sorted

                    # Pagination
                    items_per_page = 10
                    total_items = len(display_data)
                    total_pages = (
                        (total_items - 1) // items_per_page + 1
                        if total_items > 0
                        else 1
                    )

                    col1, col2, col3 = st.columns([1, 2, 1])
                    with col2:
                        page = st.selectbox("Page:", range(1, total_pages + 1))

                    start_idx = (page - 1) * items_per_page
                    end_idx = start_idx + items_per_page
                    page_data = display_data[start_idx:end_idx]

                    # Display detailed table
                    for idx, evaluation in enumerate(page_data, start_idx + 1):
                        score_val = format_eval_score(evaluation["eval_score"])
                        with st.expander(
                            f"#{idx} - {evaluation['username']} | {evaluation['department']} | Score: {score_val:.1f}%"
                        ):
                            col1, col2 = st.columns(2)

                            with col1:
                                st.write(f"**Trainee:** {evaluation['username']}")
                                st.write(f"**Industry:** {evaluation['department']}")
                                st.write(
                                    f"**Reason for Contact:** {evaluation['issue_type']}"
                                )

                            with col2:
                                st.write(f"**Tone:** {evaluation['tone']}")
                                st.write(f"**Timestamp:** {evaluation['timestamp']}")
                                st.write(f"**Score:** {score_val:.1f}%")

                            # Action buttons for individual evaluation
                            button_col1, button_col2 = st.columns(2)

                            with button_col1:
                                st.download_button(
                                    label="📄 Download Session Report",
                                    data=evaluation["excel_bytes"],
                                    file_name=f"{evaluation['username']}_{evaluation['filename']}",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    key=f"download_{idx}",
                                )

                            with button_col2:
                                if st.button(
                                    "👁️ View Report",
                                    key=f"view_supervisor_eval_{idx}",
                                    help="View detailed evaluation report",
                                ):
                                    st.session_state["view_supervisor_eval_idx"] = idx
                                    st.session_state["supervisor_eval_data"] = (
                                        evaluation
                                    )
                                    st.rerun()

                    # Summary at bottom
                    st.markdown("---")
                    st.markdown(
                        f"Showing {len(page_data)} of {total_items} evaluations"
                    )

        st.stop()

    # --- Trainee UI (original main UI) ---
    # 🚀 PERFORMANCE OPTIMIZATION: In-memory caching implemented for trainee evaluations
    # - Performance Evaluation button: Loads from GCP and caches in session state
    # - Subsequent views: Uses cached data for instant response
    # - Cache automatically cleared on logout to free memory
    # - Clears both gcs_evaluations and related performance evaluation session keys

    # --- Sidebar: Eval Summary Button ---
    if "show_eval_summary" not in st.session_state:
        st.session_state["show_eval_summary"] = False

    if st.sidebar.button("Performance Evaluation"):
        st.session_state["show_eval_summary"] = True
        st.session_state.pop("view_eval_idx", None)

        # 🚀 NEW: Load evaluations from GCS for current user
        username = st.session_state.get("username", "")
        if username:
            with st.spinner("Loading evaluations from cloud storage..."):
                gcs_manager = get_gcs_storage_manager()

                if gcs_manager.is_connected():
                    # Load all evaluations for this user from GCS
                    gcs_evaluations = gcs_manager.load_user_evaluations(username)
                    print(
                        f"[DEBUG] 📊 Loaded {len(gcs_evaluations)} evaluations from GCS for {username}"
                    )

                    # Store in session state for the performance evaluation display
                    st.session_state["gcs_evaluations"] = gcs_evaluations

                    if gcs_evaluations:
                        st.sidebar.success(
                            f"✅ Loaded {len(gcs_evaluations)} evaluation(s)"
                        )
                    else:
                        st.sidebar.info("ℹ️ No evaluations found")
                else:
                    st.sidebar.error("⚠️ Cloud storage unavailable")
                    st.session_state["gcs_evaluations"] = []
        else:
            st.sidebar.warning("⚠️ Please login first")
            st.session_state["gcs_evaluations"] = []

        st.rerun()

    # Sidebar: Back to Chat button if in summary or preview mode
    if (
        st.session_state.get("show_eval_summary", False)
        or "view_eval_idx" in st.session_state
    ):
        if st.sidebar.button("Back to Chat Session", key="sidebar_back_to_chat"):
            st.session_state.pop("show_eval_summary", None)
            st.session_state.pop("view_eval_idx", None)
            st.rerun()

    # --- Main area: Show Eval Summary Table or Excel preview if requested ---
    if st.session_state.get("show_eval_summary", False):
        # Only show summary table or Excel preview, hide chat UI
        if "view_eval_idx" not in st.session_state:
            st.title("Performance Evaluation Summary")

            # 🚀 NEW: Use GCS-loaded evaluations instead of session state
            username = st.session_state.get("username", "")
            gcs_evaluations = st.session_state.get("gcs_evaluations", [])

            if gcs_evaluations:
                st.info(
                    f"📊 Showing {len(gcs_evaluations)} evaluation(s) for {username}"
                )

                # Convert GCS evaluation data to display format
                summary_rows = []
                for display_idx, entry in enumerate(gcs_evaluations, 1):
                    summary_rows.append(
                        {
                            "S.No": display_idx,
                            "Username": entry.get("username", ""),
                            "Timestamp": entry.get("timestamp", ""),
                            "Industry": entry.get("department", ""),
                            "Reason for Contact": format_display_name(
                                entry.get("issue_type", "")
                            ),
                            "Tone": entry.get("tone", ""),
                            "Eval Score (%)": format_eval_score(
                                entry.get("eval_score", "")
                            ),
                        }
                    )
                columns = list(summary_rows[0].keys()) + ["Chat/Eval"]
                st.markdown(
                    """
                    <style>
                    .eval-table-container {
                        border-radius: 12px;
                        overflow: hidden;
                        box-shadow: 0 4px 20px rgba(0, 0, 0, 0.1);
                        margin: 1rem 0;
                        border: 1px solid #e1e8ed;
                    }
                    .eval-header {
                        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                        color: white;
                        font-weight: 600;
                        font-size: 14px;
                        padding: 1rem 0.8rem;
                        text-align: center;
                        border: none;
                        text-transform: uppercase;
                        letter-spacing: 0.5px;
                        position: relative;
                        overflow: hidden;
                    }
                    .eval-header::before {
                        content: '';
                        position: absolute;
                        top: 0;
                        left: -100%;
                        width: 100%;
                        height: 100%;
                        background: linear-gradient(90deg, transparent, rgba(255,255,255,0.2), transparent);
                        transition: left 0.5s;
                    }
                    .eval-header:hover::before {
                        left: 100%;
                    }
                    .eval-row {
                        background: #ffffff;
                        border-bottom: 1px solid #f0f4f8;
                        padding: 0.8rem;
                        transition: all 0.3s ease;
                        position: relative;
                        font-size: 14px;
                        color: #2d3748;
                    }
                    .eval-row:nth-child(even) {
                        background: #f8fafc;
                    }
                    .eval-row:hover {
                        background: linear-gradient(135deg, #667eea15 0%, #764ba215 100%);
                        transform: translateY(-1px);
                        box-shadow: 0 2px 8px rgba(102, 126, 234, 0.15);
                        border-left: 4px solid #667eea;
                        padding-left: 0.6rem;
                    }
                    .eval-row:last-child {
                        border-bottom: none;
                        border-radius: 0 0 12px 12px;
                    }
                    .eval-score {
                        font-weight: 600;
                        padding: 0.3rem 0.8rem;
                        border-radius: 20px;
                        background: linear-gradient(135deg, #48bb78 0%, #38a169 100%);
                        color: white;
                        display: inline-block;
                        min-width: 60px;
                        text-align: center;
                        box-shadow: 0 2px 4px rgba(72, 187, 120, 0.3);
                    }
                    .eval-button {
                        background: linear-gradient(135deg, #4299e1 0%, #3182ce 100%);
                        color: white;
                        border: none;
                        padding: 0.5rem 1rem;
                        border-radius: 8px;
                        font-weight: 500;
                        cursor: pointer;
                        transition: all 0.3s ease;
                        box-shadow: 0 2px 4px rgba(66, 153, 225, 0.3);
                    }
                    .eval-button:hover {
                        transform: translateY(-2px);
                        box-shadow: 0 4px 12px rgba(66, 153, 225, 0.4);
                        background: linear-gradient(135deg, #3182ce 0%, #2d70b3 100%);
                    }
                    </style>
                """,
                    unsafe_allow_html=True,
                )
                # Create table with enhanced styling
                st.markdown(
                    '<div class="eval-table-container">', unsafe_allow_html=True
                )

                header_cols = st.columns(len(columns))
                for i, col in enumerate(columns):
                    header_cols[i].markdown(
                        f'<div class="eval-header">{col}</div>', unsafe_allow_html=True
                    )

                for row_idx, row in enumerate(summary_rows):
                    row_cols = st.columns(len(columns))
                    for i, col in enumerate(columns[:-1]):
                        if col == "Eval Score (%)":
                            # Special styling for score column
                            score_value = row[col]
                            row_cols[i].markdown(
                                f'<div class="eval-row"><span class="eval-score">{score_value:.1f}%</span></div>',
                                unsafe_allow_html=True,
                            )
                        else:
                            row_cols[i].markdown(
                                f'<div class="eval-row">{row[col]}</div>',
                                unsafe_allow_html=True,
                            )
                    # View button with custom styling
                    with row_cols[-1]:
                        st.markdown('<div class="eval-row">', unsafe_allow_html=True)
                        if st.button(
                            "👁️ View",
                            key=f"view_excel_{row_idx+1}",
                            help="View detailed evaluation",
                        ):
                            st.session_state["view_eval_idx"] = (
                                row_idx  # Use the correct 0-based index
                            )
                            st.session_state["show_eval_summary"] = True
                            st.rerun()
                        st.markdown("</div>", unsafe_allow_html=True)

                st.markdown("</div>", unsafe_allow_html=True)
                st.markdown("---")
            else:
                st.info("No evaluation records found for this user/session.")
        else:
            # Excel preview page with enhanced styling
            idx = st.session_state["view_eval_idx"]
            # 🚀 NEW: Use GCS-loaded evaluations instead of filtered session state
            gcs_evaluations = st.session_state.get("gcs_evaluations", [])
            if 0 <= idx < len(gcs_evaluations):
                # Enhanced CSS for the detailed view
                st.markdown(
                    """
                    <style>
                    .detail-container {
                        background: linear-gradient(135deg, #f5f7fa 0%, #c3cfe2 100%);
                        border-radius: 15px;
                        padding: 2rem;
                        margin: 1rem 0;
                        box-shadow: 0 8px 32px rgba(0, 0, 0, 0.1);
                        border: 1px solid rgba(255, 255, 255, 0.2);
                    }
                    .detail-header {
                        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                        color: white;
                        padding: 1.5rem;
                        border-radius: 12px;
                        text-align: center;
                        margin-bottom: 2rem;
                        box-shadow: 0 4px 15px rgba(102, 126, 234, 0.3);
                    }
                    .sheet-container {
                        background: white;
                        padding: 1rem;
                        margin: 0.5rem 0;
                        border-left: 3px solid #667eea;
                    }
                    .sheet-title {
                        color: #2d3748;
                        font-size: 1.3rem;
                        font-weight: 600;
                        margin-bottom: 1rem;
                        padding-bottom: 0.5rem;
                        border-bottom: 2px solid #e2e8f0;
                        display: flex;
                        align-items: center;
                        gap: 0.5rem;
                    }
                    .chat-message {
                        background: #f8fafc;
                        border-radius: 6px;
                        padding: 0.6rem 0.8rem;
                        margin: 0.3rem 0;
                        border-left: 3px solid #4299e1;
                        font-size: 0.9rem;
                        line-height: 1.4;
                    }
                    .chat-message.user {
                        background: #e6f3ff;
                        border-left-color: #3182ce;
                    }
                    .chat-message.assistant {
                        background: #f0f9ff;
                        border-left-color: #0ea5e9;
                    }
                    .chat-role {
                        font-weight: 600;
                        font-size: 0.75rem;
                        color: #4a5568;
                        margin-bottom: 0.3rem;
                        text-transform: uppercase;
                        letter-spacing: 0.3px;
                    }
                    .behavior-score {
                        display: inline-block;
                        background: linear-gradient(135deg, #48bb78 0%, #38a169 100%);
                        color: white;
                        padding: 0.3rem 1rem;
                        border-radius: 20px;
                        font-weight: 600;
                        font-size: 0.9rem;
                        margin: 0.2rem;
                        box-shadow: 0 2px 4px rgba(72, 187, 120, 0.3);
                    }
                    .behavior-item {
                        background: #f9fafb;
                        border-radius: 4px;
                        padding: 0.4rem;
                        margin: 0.2rem 0;
                        border-left: 3px solid #e5e7eb;
                        font-size: 0.85rem;
                    }
                    .behavior-item:hover {
                        background: #f3f4f6;
                        border-color: #d1d5db;
                    }
                    .back-button {
                        background: linear-gradient(135deg, #ed8936 0%, #dd6b20 100%);
                        color: white;
                        border: none;
                        padding: 0.8rem 2rem;
                        border-radius: 25px;
                        font-weight: 600;
                        cursor: pointer;
                        transition: all 0.3s ease;
                        box-shadow: 0 4px 15px rgba(237, 137, 54, 0.3);
                        margin: 2rem 0;
                    }
                    .back-button:hover {
                        transform: translateY(-2px);
                        box-shadow: 0 6px 20px rgba(237, 137, 54, 0.4);
                    }
                    .metadata-card {
                        background: linear-gradient(135deg, #fef5e7 0%, #fed7aa 100%);
                        border-radius: 10px;
                        padding: 1rem;
                        margin: 1rem 0;
                        border-left: 4px solid #f59e0b;
                    }
                    </style>
                    """,
                    unsafe_allow_html=True,
                )

                st.markdown('<div class="detail-container">', unsafe_allow_html=True)
                st.markdown(
                    f'<div class="detail-header"><h2>📊 Detailed Evaluation Report #{idx+1}</h2></div>',
                    unsafe_allow_html=True,
                )

                # Display session metadata in a card
                eval_entry = gcs_evaluations[idx]
                st.markdown(
                    f"""
                    <div class="metadata-card">
                        <strong>📝 Session Information:</strong><br>
                        <strong>Trainee:</strong> {eval_entry.get('username', 'N/A')} |
                        <strong>Department:</strong> {eval_entry.get('department', 'N/A')} |
                        <strong>Issue Type:</strong> {eval_entry.get('issue_type', 'N/A')}<br>
                        <strong>Tone:</strong> {eval_entry.get('tone', 'N/A')} |
                        <strong>Date:</strong> {eval_entry.get('timestamp', 'N/A')} |
                        <strong>Score:</strong> <span class="behavior-score">{format_eval_score(eval_entry.get('eval_score', 'N/A')):.1f}%</span>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )

                excel_bytes = gcs_evaluations[idx]["excel_bytes"]
                with BytesIO(excel_bytes) as excel_io:
                    wb = openpyxl.load_workbook(excel_io)

                    for sheet_name in wb.sheetnames:
                        st.markdown(
                            '<div class="sheet-container">', unsafe_allow_html=True
                        )

                        # Determine sheet icon
                        if (
                            "chat" in sheet_name.lower()
                            or "conversation" in sheet_name.lower()
                        ):
                            icon = "💬"
                        elif (
                            "evaluation" in sheet_name.lower()
                            or "behavior" in sheet_name.lower()
                        ):
                            icon = "📊"
                        elif "metadata" in sheet_name.lower():
                            icon = "📋"
                        else:
                            icon = "📄"

                        st.markdown(
                            f'<div class="sheet-title">{icon} {sheet_name}</div>',
                            unsafe_allow_html=True,
                        )

                        ws = wb[sheet_name]
                        data = list(ws.values)

                        if data:
                            df = pd.DataFrame(data[1:], columns=data[0])

                            # Debug: Print sheet info
                            print(
                                f"[DEBUG] Sheet: {sheet_name}, Columns: {list(df.columns)}"
                            )
                            print(f"[DEBUG] Data shape: {df.shape}")

                            # Special formatting for different sheet types
                            if (
                                "chat" in sheet_name.lower()
                                or "conversation" in sheet_name.lower()
                            ):
                                # Enhanced chat history display
                                st.markdown("📝 **Chat Conversation:**")

                                if df.empty:
                                    st.info("No chat history found in this sheet.")
                                else:
                                    for _, row in df.iterrows():
                                        # Try different possible column names
                                        role = ""
                                        content = ""

                                        # Check for role column
                                        for col in df.columns:
                                            col_lower = str(col).lower()
                                            if "role" in col_lower:
                                                role = (
                                                    str(row[col]).lower()
                                                    if not pd.isna(row[col])
                                                    else ""
                                                )
                                                break

                                        # Check for content column
                                        for col in df.columns:
                                            col_lower = str(col).lower()
                                            if (
                                                "content" in col_lower
                                                or "message" in col_lower
                                            ):
                                                content = (
                                                    str(row[col])
                                                    if not pd.isna(row[col])
                                                    else ""
                                                )
                                                break

                                        # Skip empty rows
                                        if not content or content == "nan":
                                            continue

                                        role_icon = "👤" if role == "user" else "🤖"
                                        role_label = (
                                            "Support Representative"
                                            if role == "user"
                                            else "AI Customer"
                                        )

                                        st.markdown(
                                            f"""
                                            <div class="chat-message {role}">
                                                <div class="chat-role">{role_icon} {role_label}</div>
                                                <div>{content}</div>
                                            </div>
                                            """,
                                            unsafe_allow_html=True,
                                        )

                            elif (
                                "evaluation" in sheet_name.lower()
                                or "behavior" in sheet_name.lower()
                                or "table" in sheet_name.lower()
                            ):
                                # Enhanced evaluation table display
                                st.markdown("🎯 **Behavior Evaluation Scores:**")

                                if df.empty:
                                    st.info("No evaluation data found in this sheet.")
                                else:
                                    # Add column headers for 4-column layout
                                    st.markdown(
                                        """
                                        <div class="behavior-item" style="background: #667eea; color: white; margin-bottom: 0.5rem;">
                                            <table style="width: 100%; border-collapse: collapse; font-size: 0.85rem; table-layout: fixed;">
                                                <tr>
                                                    <th style="font-weight: 600; padding: 0.6rem 0.5rem; text-align: center; width: 20%; border-right: 1px solid rgba(255,255,255,0.3);">
                                                        📋 Behavior
                                                    </th>
                                                    <th style="font-weight: 600; padding: 0.6rem 0.5rem; text-align: center; width: 15%; border-right: 1px solid rgba(255,255,255,0.3);">
                                                        🎯 Score
                                                    </th>
                                                    <th style="font-weight: 600; padding: 0.6rem 0.5rem; text-align: center; width: 30%; border-right: 1px solid rgba(255,255,255,0.3);">
                                                        💬 Citation
                                                    </th>
                                                    <th style="font-weight: 600; padding: 0.6rem 0.5rem; text-align: center; width: 35%;">
                                                        📝 Feedback
                                                    </th>
                                                </tr>
                                            </table>
                                        </div>
                                        """,
                                        unsafe_allow_html=True,
                                    )

                                    # Debug: Show all available columns
                                    print(
                                        f"[DEBUG] Available columns in {sheet_name}: {list(df.columns)}"
                                    )
                                    print(f"[DEBUG] First few rows: {df.head()}")

                                    for _, row in df.iterrows():
                                        # Try different possible column names with more flexibility
                                        behavior = ""
                                        criteria = ""
                                        max_score = ""
                                        score = ""
                                        feedback = ""
                                        reasoning = ""
                                        citation = ""

                                        # Print all row data for debugging
                                        print(f"[DEBUG] Row data: {dict(row)}")

                                        # More comprehensive column matching
                                        for col in df.columns:
                                            col_str = str(col).strip()
                                            col_lower = col_str.lower()
                                            row_value = (
                                                str(row[col]).strip()
                                                if not pd.isna(row[col])
                                                else ""
                                            )

                                            # Check for behavior column
                                            if not behavior and (
                                                "behavior" in col_lower
                                                or "behaviour" in col_lower
                                            ):
                                                behavior = row_value

                                            # Check for criteria column
                                            elif not criteria and (
                                                "criteria" in col_lower
                                                or "criterion" in col_lower
                                            ):
                                                criteria = row_value

                                            # Check for max score column
                                            elif not max_score and (
                                                "max" in col_lower
                                                and "score" in col_lower
                                            ):
                                                max_score = row_value

                                            # Check for score column
                                            elif not score and (
                                                col_lower == "score"
                                                or (
                                                    col_lower.endswith("score")
                                                    and "max" not in col_lower
                                                )
                                            ):
                                                score = row_value

                                            # Check for feedback/reasoning column - prioritize "Reason" column
                                            elif not feedback and (
                                                col_lower == "reason"
                                                or col_lower == "reasons"
                                                or any(
                                                    pattern in col_lower
                                                    for pattern in [
                                                        "feedback",
                                                        "comment",
                                                        "remarks",
                                                        "notes",
                                                        "suggestion",
                                                        "improvement",
                                                        "reasoning",
                                                    ]
                                                )
                                            ):
                                                feedback = row_value

                                            # Check for citation column
                                            elif not citation and any(
                                                pattern in col_lower
                                                for pattern in [
                                                    "citation",
                                                    "cite",
                                                    "reference",
                                                    "quote",
                                                    "example",
                                                ]
                                            ):
                                                citation = row_value

                                        # Fallback logic
                                        if not behavior and len(df.columns) > 0:
                                            behavior = (
                                                str(row[df.columns[0]]).strip()
                                                if not pd.isna(row[df.columns[0]])
                                                else ""
                                            )

                                        # Try to find score in any numeric column
                                        if not score:
                                            for col in df.columns:
                                                try:
                                                    val = str(row[col]).strip()
                                                    if (
                                                        val
                                                        and val != "nan"
                                                        and val.replace(".", "")
                                                        .replace(",", "")
                                                        .isdigit()
                                                    ):
                                                        score = val
                                                        break
                                                except:
                                                    continue

                                        # Try to find feedback/reasoning in longer text columns - prioritize "Reason" column
                                        if not feedback or feedback in [
                                            "nan",
                                            "None",
                                            "",
                                            "No feedback provided",
                                            "Performance evaluation criteria",
                                        ]:
                                            # First, check specifically for "Reason" or "Reasons" columns
                                            for col in df.columns:
                                                col_lower = str(col).lower()
                                                if (
                                                    col_lower == "reason"
                                                    or col_lower == "reasons"
                                                ):
                                                    val = (
                                                        str(row[col]).strip()
                                                        if not pd.isna(row[col])
                                                        else ""
                                                    )
                                                    if val and val != "nan":
                                                        feedback = val
                                                        break

                                            # If still no feedback, try other text columns
                                            if not feedback or feedback in [
                                                "nan",
                                                "None",
                                                "",
                                            ]:
                                                for col in df.columns:
                                                    col_lower = str(col).lower()
                                                    if any(
                                                        skip in col_lower
                                                        for skip in [
                                                            "behavior",
                                                            "behaviour",
                                                            "criteria",
                                                            "score",
                                                            "max",
                                                        ]
                                                    ):
                                                        continue

                                                    val = (
                                                        str(row[col]).strip()
                                                        if not pd.isna(row[col])
                                                        else ""
                                                    )
                                                    if (
                                                        val
                                                        and val != "nan"
                                                        and len(val) > 15
                                                    ):  # Look for substantial text
                                                        feedback = val
                                                        break

                                        # Skip completely empty rows
                                        if not behavior or behavior in [
                                            "nan",
                                            "None",
                                            "",
                                        ]:
                                            continue

                                        # Clean up values
                                        max_score = (
                                            max_score
                                            if max_score
                                            and max_score not in ["nan", "None", ""]
                                            else "5"
                                        )
                                        score = (
                                            score
                                            if score
                                            and score not in ["nan", "None", ""]
                                            else "0"
                                        )
                                        feedback = (
                                            feedback
                                            if feedback
                                            and feedback
                                            not in [
                                                "nan",
                                                "None",
                                                "",
                                                "Performance evaluation criteria",
                                            ]
                                            else ""
                                        )
                                        citation = (
                                            citation
                                            if citation
                                            and citation not in ["nan", "None", ""]
                                            else ""
                                        )

                                        # Debug print the extracted values
                                        print(
                                            f"[DEBUG] Extracted - Behavior: {behavior}, Score: {score}, Feedback: {feedback[:100]}..., Citation: {citation[:50]}..."
                                        )

                                        # Create 4-column table layout with full text wrapping (no truncation)
                                        # Function to clean and prepare text for display
                                        def clean_text(text):
                                            if not text or text in ["nan", "None", ""]:
                                                return "<em>No data provided</em>"
                                            return (
                                                str(text).strip().replace("\n", "<br>")
                                            )

                                        clean_behavior = clean_text(behavior)
                                        clean_citation = clean_text(citation)
                                        clean_feedback = clean_text(feedback)

                                        st.markdown(
                                            f"""
                                            <div class="behavior-item">
                                                <table style="width: 100%; border-collapse: collapse; font-size: 0.85rem; table-layout: fixed;">
                                                    <tr style="vertical-align: top;">
                                                        <td style="font-weight: 600; color: #2d3748; padding: 0.6rem 0.5rem; border-right: 1px solid #e5e7eb; width: 20%; word-wrap: break-word; overflow-wrap: break-word; white-space: normal; line-height: 1.4;">
                                                            🎯 {clean_behavior}
                                                        </td>
                                                        <td style="text-align: center; padding: 0.6rem 0.5rem; border-right: 1px solid #e5e7eb; width: 15%; word-wrap: break-word; vertical-align: middle;">
                                                            <span class="behavior-score">{score}/{max_score}</span>
                                                        </td>
                                                        <td style="padding: 0.6rem 0.5rem; border-right: 1px solid #e5e7eb; width: 30%; word-wrap: break-word; overflow-wrap: break-word; white-space: normal; font-style: italic; color: #4a5568; line-height: 1.4;">
                                                            💬 {clean_citation}
                                                        </td>
                                                        <td style="padding: 0.6rem 0.5rem; width: 35%; word-wrap: break-word; overflow-wrap: break-word; white-space: normal; line-height: 1.4;">
                                                            {clean_feedback}
                                                        </td>
                                                    </tr>
                                                </table>
                                            </div>
                                            """,
                                            unsafe_allow_html=True,
                                        )
                            else:
                                # Regular table display for other sheets
                                st.markdown(f"📋 **{sheet_name} Data:**")
                                st.dataframe(df, use_container_width=True)
                        else:
                            st.info("No data in this sheet.")

                        st.markdown("</div>", unsafe_allow_html=True)

                st.markdown("</div>", unsafe_allow_html=True)

                # Enhanced back button
                st.markdown("---")
                col1, col2, col3 = st.columns([1, 1, 1])
                with col2:
                    if st.button(
                        "⬅️ Back to Summary",
                        key="back_to_summary_btn",
                        help="Return to evaluation summary",
                    ):
                        st.session_state.pop("view_eval_idx", None)
                        st.session_state["show_eval_summary"] = True
                        st.rerun()
            st.stop()
        # Do not show chat UI when summary is active
        # --- Plots below summary table ---
        st.markdown("---")
        # ---- Eval Score Trend ----
        st.markdown("### 📈 Evaluation Score Trend (Line Graph)")
        # Reverse the order for chronological display (oldest to newest)
        gcs_evaluations = st.session_state.get("gcs_evaluations", [])
        scores = [
            float(entry.get("eval_score", "0").replace("%", ""))
            for entry in reversed(gcs_evaluations)
        ]
        if scores:
            # Create a clean, modern line plot with seaborn
            plot_df = pd.DataFrame(
                {"Trial": np.arange(1, len(scores) + 1), "Score": scores}
            )

            # Set the style and color palette
            sns.set_style("whitegrid")
            plt.style.use("seaborn-v0_8")

            # Calculate dynamic figure size
            fig_width = max(8, min(12, len(scores) * 1.2))
            fig_height = 5

            fig, ax = plt.subplots(figsize=(fig_width, fig_height))

            # Create the line plot with enhanced styling
            line_plot = sns.lineplot(
                data=plot_df,
                x="Trial",
                y="Score",
                marker="o",
                markersize=8,
                linewidth=3,
                color="#2E86AB",
                ax=ax,
            )

            # Fill area under the curve for better visual appeal
            ax.fill_between(
                plot_df["Trial"], plot_df["Score"], alpha=0.3, color="#2E86AB"
            )

            # Add value annotations on each point
            for x, y in zip(plot_df["Trial"], plot_df["Score"]):
                ax.annotate(
                    f"{y:.1f}%",
                    (x, y),
                    textcoords="offset points",
                    xytext=(0, 10),
                    ha="center",
                    fontsize=10,
                    fontweight="bold",
                    bbox=dict(boxstyle="round,pad=0.3", facecolor="white", alpha=0.8),
                )

            # Customize the plot
            ax.set_xlabel("Training Session", fontsize=12, fontweight="bold")
            ax.set_ylabel("Evaluation Score (%)", fontsize=12, fontweight="bold")
            ax.set_title(
                "Performance Trend Across Training Sessions",
                fontsize=14,
                fontweight="bold",
                pad=20,
            )

            # Set axis limits and ticks
            ax.set_xlim(0.5, len(scores) + 0.5)
            ax.set_ylim(0, 105)
            ax.set_xticks(plot_df["Trial"])
            ax.set_yticks(np.arange(0, 101, 10))

            # Customize grid
            ax.grid(True, linestyle="--", alpha=0.7, color="gray")
            ax.set_axisbelow(True)

            # Add a horizontal line for average score
            avg_score = np.mean(scores)
            ax.axhline(y=avg_score, color="red", linestyle="--", alpha=0.7, linewidth=2)
            ax.text(
                len(scores),
                avg_score,
                f"Avg: {avg_score:.1f}%",
                verticalalignment="bottom",
                fontsize=10,
                color="red",
                fontweight="bold",
            )

            # Improve tick labels
            ax.tick_params(axis="both", which="major", labelsize=10)

            plt.tight_layout()
            st.pyplot(fig)
            plt.close()
        else:
            st.info("No evaluation scores to plot.")

        st.markdown("---")
        # ---- Behavior Scores Histogram ----
        st.markdown("### 🏆 Behavior Scores (Average)")

        behavior_totals = {}
        behavior_counts = {}
        total_sessions = len(gcs_evaluations)

        for entry in gcs_evaluations:
            eb = entry.get("excel_bytes")
            if eb:
                with BytesIO(eb) as io:
                    wb = openpyxl.load_workbook(io, data_only=True)
                    if "Evaluation Table" in wb.sheetnames:
                        ws = wb["Evaluation Table"]
                        for r in ws.iter_rows(min_row=2, values_only=True):
                            b, _, mx, sc, *_ = r
                            try:
                                sc = float(sc)
                                if b:
                                    behavior_totals[b] = behavior_totals.get(b, 0) + sc
                                    behavior_counts[b] = behavior_counts.get(b, 0) + 1
                            except:
                                continue

        # Display session summary before the graph
        st.info(
            f"📊 **Analysis Summary:** {total_sessions} total training sessions completed"
        )

        if behavior_totals:
            behaviors = list(behavior_totals.keys())
            # Calculate averages instead of totals
            avg_scores = [behavior_totals[b] / behavior_counts[b] for b in behaviors]
            session_counts = [behavior_counts[b] for b in behaviors]

            plot_df = pd.DataFrame(
                {
                    "Behavior": behaviors,
                    "Average Score": avg_scores,
                    "Sessions": session_counts,
                }
            )

            # Sort by average score for better visualization
            plot_df = plot_df.sort_values("Average Score", ascending=True)

            # Set the style
            sns.set_style("whitegrid")
            plt.style.use("seaborn-v0_8")

            # Calculate dynamic figure size based on number of behaviors
            fig_height = max(6, len(behaviors) * 0.6)
            fig_width = 12

            fig, ax = plt.subplots(figsize=(fig_width, fig_height))

            # Create a beautiful horizontal bar plot
            bars = sns.barplot(
                data=plot_df,
                y="Behavior",
                x="Average Score",
                ax=ax,
                palette="viridis",
                orient="h",
            )

            # Add value labels on the bars with session count
            for i, (behavior, avg_score, session_count) in enumerate(
                zip(plot_df["Behavior"], plot_df["Average Score"], plot_df["Sessions"])
            ):
                ax.text(
                    avg_score + max(avg_scores) * 0.01,
                    i,
                    f"{avg_score:.1f} ({session_count} sessions)",
                    va="center",
                    ha="left",
                    fontsize=10,
                    fontweight="bold",
                )

            # Customize the plot
            ax.set_xlabel("Average Score", fontsize=12, fontweight="bold")
            ax.set_ylabel("Behavior", fontsize=12, fontweight="bold")
            ax.set_title(
                "Average Behavior Performance Scores",
                fontsize=14,
                fontweight="bold",
                pad=20,
            )

            # Improve the y-axis labels (behavior names)
            ax.tick_params(axis="y", labelsize=10)
            ax.tick_params(axis="x", labelsize=10)

            # Set x-axis limits with some padding
            ax.set_xlim(0, max(avg_scores) * 1.15)

            # Customize grid
            ax.grid(True, linestyle="--", alpha=0.7, color="gray", axis="x")
            ax.set_axisbelow(True)

            # Add a subtle background color gradient
            for i, bar in enumerate(bars.patches):
                bar.set_edgecolor("white")
                bar.set_linewidth(1)

            plt.tight_layout()
            st.pyplot(fig)
            plt.close()
        else:
            st.info("No behavior scores to plot.")

        st.stop()
    st.set_page_config(
        page_title="AI Customer Support Trainer", page_icon="🎯", layout="wide"
    )

    st.title("🎯 AI Customer Support Trainer")
    st.markdown(
        "<h3 style='font-size: 24px; font-weight: 600; color: #2E86AB; margin-bottom: 30px;'>Practice your customer support skills with AI-powered customer training simulator</h3>",
        unsafe_allow_html=True,
    )

    # Associate Login Banner
    current_user = st.session_state.get("username", "Unknown User")
    st.markdown(
        f"<div style='background: linear-gradient(135deg, #e3f2fd 0%, #bbdefb 100%); color: #1565c0; padding: 6px 15px; border-radius: 8px; margin-bottom: 15px; text-align: center;'>"
        f"<h5 style='margin: 0; color: #1565c0; font-size: 16px;'>👤 Associate: {current_user} Dashboard</h5>"
        f"</div>",
        unsafe_allow_html=True,
    )

    # Simplified United Airlines departments - no need for dynamic loading
    departments = [
        "Baggage Handling",
        "Boarding and Seating",
        "Flight Changes and Cancellations",
        "Mileageplus Rewards",
        "Pets and Travel",
        "Special Services",
        "Ticket Sales and Pricing",
    ]
    # Initialize trainer in session state
    # Initialize trainer in session state
    if "trainer" not in st.session_state:
        st.session_state.trainer = CustomerSupportTrainer()

    # Sidebar controls
    # Initialize reset counter for assigned simulation dropdown
    if "assigned_dropdown_reset" not in st.session_state:
        st.session_state["assigned_dropdown_reset"] = 0

    with st.sidebar:
        if st.button("🔄 Start New Session"):
            st.session_state.trainer.reset_conversation()
            st.session_state.messages = []
            st.session_state.trainer.conversation_started = False
            # Clear chat history for current user in user_data
            username = st.session_state.get("username", "")
            if username:
                if "user_data" not in st.session_state:
                    st.session_state.user_data = {}
                if username not in st.session_state.user_data:
                    st.session_state.user_data[username] = {}
                st.session_state.user_data[username]["chat_history"] = []
                st.session_state.chat_history = []
            # NEW: Also clear chat_history from session state on next login
            st.session_state["clear_chat_on_login"] = True
            # Bump reset counter so dropdowns reset to defaults on rerender
            st.session_state["assigned_dropdown_reset"] += 1
            # Reset scenario dropdown-related session values to defaults
            for k in ("scenario_type", "issue_type", "customer_tone"):
                if k in st.session_state:
                    del st.session_state[k]

            # 🚀 NEW: Clear all policy-related session data to prevent interference
            policy_keys = [
                "policy_scenarios",
                "current_policy_session",
                "policy_switch_scenario",
                "selected_policies",
                "scenario_chunks",
                "policy_context",
                "current_scenario",
            ]
            for key in policy_keys:
                if key in st.session_state:
                    if key == "policy_scenarios":
                        st.session_state[key] = []  # Reset to empty list
                    elif key == "current_policy_session":
                        st.session_state[key] = None  # Reset to None
                    else:
                        del st.session_state[key]  # Delete completely

            # Reset evaluate session button for new session
            st.session_state.evaluate_session_clicked = False

            st.success("New training session started!")

        st.markdown("---")

        # --- SUPERVISOR ASSIGNED SIMULATIONS DROPDOWN ---
        username = st.session_state.get("username", "")
        assign_list = cached_supervisor_assignments(username)

        # Prepare dropdown options
        assign_options = ["-- Select Assigned Simulation --"] + sorted(
            set(s for s in assign_list if s)
        )

        selected_assignment = st.selectbox(
            "Assigned Simulations (from Supervisor)",
            assign_options,
            key=f"assigned_simulation_dropdown_{st.session_state['assigned_dropdown_reset']}",
        )

        # Refresh button below the dropdown
        if st.button("🔄 Refresh Assignments"):
            cached_supervisor_assignments.clear()
            st.rerun()

        # Parse selected assignment into existing session variables (handle hyphens safely)
        import re

        if (
            selected_assignment
            and selected_assignment != "-- Select Assigned Simulation --"
        ):
            # Store the full simulation string for later use
            st.session_state["assigned_simulation"] = selected_assignment
            st.session_state["assigned_simulation_dropdown"] = selected_assignment

            parts = [
                p.strip() for p in re.split(r"\s*-\s*", selected_assignment, maxsplit=2)
            ]
            if len(parts) >= 1 and parts[0]:
                st.session_state["scenario_type"] = parts[0]
            if len(parts) >= 2 and parts[1]:
                st.session_state["issue_type"] = parts[1]
            if len(parts) >= 3 and parts[2]:
                st.session_state["customer_tone"] = parts[2]

        # Get values from session state (populated by assignment parsing above)
        scenario_type = st.session_state.get("scenario_type", "")
        issue_type = st.session_state.get("issue_type", "")
        customer_tone = st.session_state.get("customer_tone", "")

        st.markdown("---")
        st.markdown(f"**Current Scenario:** {scenario_type}")
        st.markdown(f"**Reason for Contact:** {issue_type}")
        st.markdown(f"**Customer Tone:** {customer_tone}")

        # Export chat to Excel download button using buffer memory
        st.markdown("---")
        chat_history = st.session_state.trainer.get_conversation_history()
        rows = []
        for msg in chat_history:
            if isinstance(msg, HumanMessage):
                rows.append({"role": "user", "content": msg.content})
            elif isinstance(msg, AIMessage):
                rows.append({"role": "assistant", "content": msg.content})
        if rows:
            df = pd.DataFrame(rows)
            output = BytesIO()
            df.to_excel(output, index=False, engine="openpyxl")
            output.seek(0)
        else:
            st.info("No chat history to export.")

        # Note: Chat history download removed - data is preserved in evaluation Excel files

        # Logout button at bottom of sidebar
        st.markdown("---")
        if st.button("🚪 Logout", key="logout_btn", type="secondary"):
            # 🚀 COMPREHENSIVE LOGOUT CLEANUP:
            # - Login credentials
            # - Cached evaluation data
            # - Performance evaluation UI data
            # - Policy-related session data (same as Start New Session)
            # - Conversation memory and chat history
            print(f"[DEBUG] User {st.session_state.get('username', '')} logging out")
            st.session_state.logged_in = False
            st.session_state.username = ""
            st.session_state.role = ""

            # 🚀 Clear trainee cached evaluation data to free memory
            if "gcs_evaluations" in st.session_state:
                cached_count = len(st.session_state["gcs_evaluations"])
                del st.session_state["gcs_evaluations"]
                print(
                    f"[DEBUG] 💾 Cleared {cached_count} cached evaluations from trainee memory"
                )

            # Clear other performance evaluation session data
            performance_keys = [
                "show_eval_summary",
                "view_eval_idx",
                "supervisor_eval_data",
                "view_supervisor_eval_idx",
                "evaluate_session_clicked",  # Clear evaluate session button state
            ]
            for key in performance_keys:
                if key in st.session_state:
                    del st.session_state[key]

            # 🚀 Clear policy-related session data (same as Start New Session)
            policy_keys = [
                "policy_scenarios",
                "current_policy_session",
                "policy_switch_scenario",
                "selected_policies",
                "scenario_chunks",
                "policy_context",
                "current_scenario",
                "messages",
            ]
            for key in policy_keys:
                if key in st.session_state:
                    if key == "policy_scenarios":
                        st.session_state[key] = []  # Reset to empty list
                    elif key == "current_policy_session":
                        st.session_state[key] = None  # Reset to None
                    else:
                        del st.session_state[key]  # Delete completely

            #  Clear GCS storage manager cache - FIXES THE ORIGINAL ISSUE!
            # This ensures fresh data loading on next login
            try:
                get_gcs_storage_manager.clear()
            except Exception as e:
                print(f"[DEBUG] ⚠️ Could not clear GCS cache: {e}")

            st.rerun()
    # --- Main area: Show Excel preview if requested ---

    # Main chat interface

    # Display conversation history from buffer memory
    # Use session chat_history for current user
    chat_history = st.session_state.get("chat_history", [])
    for msg in chat_history:
        if isinstance(msg, HumanMessage):
            with st.chat_message("user"):
                st.markdown(msg.content)
        elif isinstance(msg, AIMessage):
            with st.chat_message("assistant"):
                st.markdown(msg.content)

    # Show initial customer message if conversation hasn't started
    # NEW LOGIC: Require scenario, issue, and tone before starting chat
    required_ready = all(
        bool(str(st.session_state.get(k, "")).strip())
        for k in ("scenario_type", "issue_type", "customer_tone")
    )
    if not st.session_state.trainer.conversation_started:
        if not required_ready:
            st.info(
                "Please select an assigned simulation (or set Scenario, Issue Type, and Customer Tone) in the sidebar before starting the conversation."
            )
        else:
            st.info("Support Rep, please start the conversation.")
    # Only set conversation_started after first support rep input

    # Support rep input (only when required selections are ready)
    support_input = None
    if required_ready:
        support_input = st.chat_input(
            "Type your response as a support representative..."
        )
    if support_input:
        # Always show support rep message in UI
        with st.chat_message("user"):
            st.markdown(support_input)

        st.session_state.trainer.conversation_started = True

        # Load policy documents
        policy_db = get_policy_faiss_db()
        all_matching_docs = []

        if policy_db:
            try:
                # Filter documents by theme ONLY - load only matching theme documents
                for doc_id, doc in policy_db.docstore._dict.items():
                    doc_theme = doc.metadata.get("theme", "")
                    if (
                        doc_theme == issue_type
                    ):  # Only add documents that match the selected theme
                        all_matching_docs.append(doc)

                print(
                    f"[DEBUG] Filtered {len(all_matching_docs)} documents for theme '{issue_type}'"
                )
            except Exception as e:
                print(f"[DEBUG] Error retrieving policy documents: {e}")
                st.warning(f"Error retrieving policy documents: {str(e)}")

        # Use policy-based response
        if all_matching_docs:
            customer_response = st.session_state.trainer.gen_cus_policy_response(
                support_input,
                all_matching_docs,
                customer_tone,
                scenario_type,
                issue_type,
            )
        else:
            # Fallback when no policies available
            customer_response = "I understand your inquiry. Let me connect you with a supervisor who can better assist you with this matter."

        # Always show AI customer response in UI
        with st.chat_message("assistant"):
            with st.spinner("Customer is typing..."):
                st.markdown(customer_response)

        # Save chat history to session for current user
        st.session_state.chat_history = (
            st.session_state.trainer.get_conversation_history()
        )

        if not st.session_state.trainer.conversation_started and chat_history:
            st.info("Conversation ended. You can start a new session from the sidebar.")

    # Exit conversation button
    if chat_history and len(chat_history) > 1:
        st.markdown("---")
        col1, col2, col3 = st.columns([1, 1, 1])
        with col2:
            if st.button("🚪 End Training Session", type="primary"):
                st.success(
                    "Training session ended! Great job practicing your support skills."
                )
                st.markdown("**Session Summary:**")
                st.markdown(f"• Scenario: {scenario_type}")
                st.markdown(f"• Issue Type: {issue_type}")
                st.markdown(f"• Customer Tone: {customer_tone}")
                st.markdown(f"• Total exchanges: {len(chat_history)//2}")
                st.markdown("• You can start a new session from the sidebar")

    # Show evaluation button when conversation ended and chat history exists
    if not st.session_state.trainer.conversation_started and chat_history:
        st.markdown("---")
        # Check if evaluate session button has been clicked
        if st.session_state.evaluate_session_clicked:
            st.info(
                "✅ Session already evaluated. Start a new training session to evaluate again."
            )
        else:
            # Disable button if already clicked to prevent duplicate evaluations
            button_disabled = st.session_state.evaluate_session_clicked
            if st.button(
                "📝 Evaluate Session", type="primary", disabled=button_disabled
            ):
                # Mark as clicked to disable for rest of session
                st.session_state.evaluate_session_clicked = True
                # Immediately refresh the page to show the "already evaluated" message

                with st.spinner("Evaluating session with LLM..."):
                    # Load evaluation metrics vector store
                    eval_vector_store_path = "faiss_store_eval_metrics"

                    eval_embedder = EvalEmbeddings(
                        model="text-embedding-3-large",
                        api_key="83265f364fe844f298b2d4f8a5a39426",
                        openai_api_type="azure",
                        azure_endpoint="https://genai-demos.openai.azure.com/",
                        api_version="2024-02-15-preview",
                    )
                    eval_db = EvalFAISS.load_local(
                        eval_vector_store_path,
                        eval_embedder,
                        allow_dangerous_deserialization=True,
                    )
                    selected_domain = st.session_state.get("scenario_type", None)

                    # Get policy documents used for scenario creation from session state
                    policy_documents_text = ""
                    if (
                        "current_policy_session" in st.session_state
                        and st.session_state.current_policy_session is not None
                    ):
                        selected_policies = st.session_state.current_policy_session.get(
                            "selected_policies", []
                        )
                        if selected_policies:
                            policy_documents_text = (
                                "\n\nPOLICY DOCUMENTS USED FOR SCENARIO:\n"
                            )
                            for i, doc in enumerate(selected_policies, 1):
                                policy_documents_text += (
                                    f"Policy Document {i}:\n{doc.page_content}\n\n"
                                )
                            print(
                                f"[DEBUG] Found {len(selected_policies)} policy documents from session state"
                            )
                        else:
                            print("[DEBUG] No policy documents found in session state")
                    else:
                        print(
                            "[DEBUG] No current_policy_session found in session state"
                        )

                    # Map scenario types to evaluation domains
                    domain_for_policy = (
                        str(selected_domain).lower() if selected_domain else ""
                    )
                    # Map scenario types to correct evaluation domains
                    if "retail" in domain_for_policy:
                        domain_for_policy = "retail"
                    elif "air" in domain_for_policy or "airline" in domain_for_policy:
                        domain_for_policy = "airlines"

                    all_domains = set()
                    for doc in eval_db.docstore._dict.values():
                        dom = doc.metadata.get("Domain", "")
                        all_domains.add(dom)
                    print(
                        f"[DEBUG] Available domains in eval_db: {sorted(all_domains)}"
                    )
                    print(
                        f"[DEBUG] Selected domain for policy: '{domain_for_policy}' (original: '{selected_domain}')"
                    )
                    eval_metrics = []
                    for doc in eval_db.docstore._dict.values():
                        if doc.metadata.get("Domain", "").lower() == domain_for_policy:
                            eval_metrics.append(doc)
                    if not eval_metrics:
                        print(
                            f"[WARNING] No metrics found for domain '{selected_domain}'. Check for case/whitespace mismatches."
                        )
                        st.warning(
                            f"No evaluation metrics found for domain '{selected_domain}'. Please check your selection and vector store."
                        )

                    # Prepare chat history for prompt
                    history_text = ""
                    for message in chat_history:
                        if hasattr(message, "role"):
                            if message.role == "user":
                                history_text += f"Support Rep: {message.content}\n"
                            elif message.role == "assistant":
                                history_text += f"Customer: {message.content}\n"
                        else:
                            # Fallback for HumanMessage/AIMessage
                            if message.__class__.__name__ == "HumanMessage":
                                history_text += f"Support Rep: {message.content}\n"
                            elif message.__class__.__name__ == "AIMessage":
                                history_text += f"Customer: {message.content}\n"

                    # Prepare metrics for prompt
                    metrics_text = ""
                    behaviour_list = []
                    for idx, metric in enumerate(eval_metrics, 1):
                        # Extract behaviour and definition from page_content (assume format: 'Behaviour: ...\nDefinition: ...')
                        page_lines = metric.page_content.split("\n")
                        behaviour = ""
                        definition = ""
                        for line in page_lines:
                            if line.lower().startswith("behaviour:"):
                                behaviour = line.split(":", 1)[1].strip()
                            elif line.lower().startswith("definition:"):
                                definition = line.split(":", 1)[1].strip()
                        if behaviour:
                            behaviour_list.append(behaviour)
                        score = metric.metadata.get(
                            "Score", metric.metadata.get("score", "")
                        )
                        mandatory = metric.metadata.get(
                            "Mandatory", metric.metadata.get("mandatory", "")
                        )
                        description = ""
                        if domain_for_policy == "Airlines":
                            pass

                        else:
                            description = metric.metadata.get(
                                "Description", metric.metadata.get("description", "")
                            )
                        metrics_text += f"Metric {idx}:\nBehaviour: {behaviour}\nDefinition: {definition}\nRecommended Score: {score}\nMandatory: {mandatory}\nDescription: {description}\n\n"
                    print(f"____behav list ______ {behaviour_list}")

                    eval_prompt = ""
                    if domain_for_policy == "retail":
                        eval_prompt = f"""
                     You are an expert evaluator for customer support conversations.

                     SCENARIO DOMAIN: {selected_domain}

                     EVALUATION METRICS:
                     {metrics_text}
                    POLICY DOCUMENTS (Use to evaluate the metrics):
                     {policy_documents_text}

                     CHAT HISTORY:
                     {history_text}

                     STRICT INSTRUCTIONS:
                        1. You MUST evaluate each behavior separately using the EXACT behavior names provided below:
                        {behaviour_list}

                        2. BEHAVIOR EVALUATION LOGIC - Follow this decision process for each behavior:
                            - Check if behavior is MANDATORY in the metadata
                            - If mandatory = TRUE: ALWAYS evaluate this behavior (score 0 to max)
                            - If mandatory = FALSE: Read the description and analyze if it applies to this conversation
                            - If description conditions don't match the conversation context, mark as "Not Applicable"

                        3. POLICY COMPLIANCE CHECK:
                            - For behaviors "Resolution", "Policy/Product/Service Knowledge", and "Process":
                            - Check if the support rep's responses followed the policy documents provided above
                            - Include policy compliance in your scoring and reasoning

                       4. CITATION LOGIC - VERY IMPORTANT:
                         - In the Citation column, include ONLY the ACTUAL CHAT MESSAGE CONTENT from the support representative that you used to evaluate this behavior
                         - DO NOT include any policy document content or references in the citation
                         - DO NOT use generic references like "Chat #2" or "Message 3"
                         - Copy the exact support representative message(s) that demonstrate (or fail to demonstrate) the behavior
                         - ONLY cite content from the CHAT HISTORY section, never from the POLICY DOCUMENTS section
                         - If you used the entire conversation history for evaluation, write "Entire conversation history used"
                         - If you used multiple chat messages, include all relevant chat message content
                         - Remember: Citations should ONLY contain what the support rep said, not policy information

                        5. SCORING LOGIC:
                         - Each behavior has a MAXIMUM possible score from the metadata
                         - IF behavior is FULLY demonstrated: Give the maximum score from metadata
                         - IF behavior is PARTIALLY demonstrated: Give 75% of maximum score
                         - IF behavior is NOT present: Give 0
                         - IF behavior is NOT APPLICABLE: Mark as "N/A"

                        6. REASON COLUMN REQUIREMENTS - CRITICAL:
                         - DO NOT just restate the definition in the Reason column
                         - Provide SPECIFIC analysis of what the support rep did or failed to do
                         - Explain HOW their actions/responses align or don't align with the behavior definition
                         - Reference specific policy compliance or violations when relevant
                         - For FULL scores: Explain what the rep did well that earned the maximum score
                         - For PARTIAL scores: Explain what was missing or could be improved (why not full score)
                         - For ZERO scores: Explain what specific behavior was absent or done incorrectly
                         - For N/A scores: Explain why this behavior doesn't apply to this conversation
                        

                        7. OUTPUT FORMAT - Use this exact table structure:

                         | Behavior | Definition | Max Score | Awarded Score | Citation | Reason |
                            |----------|------------|-----------|---------------|----------|-----------|
                        | [Exact name] | [Definition] | [From metadata] | [Your score] | [Reference to chat history] | [Reason] |

                        8. DO NOT:
                         - Group behaviors together
                         - Create new behavior names
                         - Invent evidence not in chat history
                         - Change the behavior names provided

                     """
                    else:
                        eval_prompt = f"""
                        you are an Expert evaluator for customer support conversation
                        SCENARIO DOMAIN: 
                        {selected_domain}
                        Polices Documents (Use to evaluate the metrics):
                        {policy_documents_text}
                        EVALUATION METRICS:
                        {metrics_text}
                        CHAT HISTORY:
                        {history_text}
                        STRICT INSTRUCTIONS:
                            1. You MUST evaluate each behavior separately using the EXACT behavior names provided below:
                            {behaviour_list}
                            2. BEHAVIOR EVALUATION LOGIC - Follow this decision process for each behavior:
                                -check for the definition of each behaviours in the metrics {metrics_text}
                                -If the definition matches the conversation context, evaluate the behavior:
                                      1.check the score from the {metrics_text}, based on that you can assign a score to the behaviour.
                                      2. If the behaviour is MANDATORY in the metadata, ALWAYS evaluate this behaviour else evaluate if the behaviour matches context of the conversation.
                                      3. If the behaviour does not match the context of the conversation, mark as "Not Applicable"
                                      4. if the behaviour is fully demonstrated, as per the {metrics_text} assign the score from the any of three mentioned criteria.
                            3. POLICY COMPLIANCE CHECK:
                                - Check if the support rep's responses followed the policy documents provided above
                                - Include policy compliance in your  reasoning

                            4. CITATION LOGIC - VERY IMPORTANT:
                                - In the Citation column, include ONLY the ACTUAL CHAT MESSAGE CONTENT from the support representative that you used to evaluate this behavior
                                - DO NOT include any policy document content or references in the citation
                                - DO NOT use generic references like "Chat #2" or "Message 3"
                                - Copy the exact support representative message(s) that demonstrate (or fail to demonstrate) the behavior
                                - ONLY cite content from the CHAT HISTORY section, never from the POLICY DOCUMENTS section
                                - If you used the entire conversation history for evaluation, write "Entire conversation history used"
                                - If you used multiple chat messages, include all relevant chat message content
                                - Remember: Citations should ONLY contain what the support rep said, not policy information
                            5. SCORING LOGIC:
                                - Each behavior has three categories of score as per the {metrics_text}
                                - IF behavior is FULLY demonstrated: Then it is  Meets/Delivered and assign score 10
                                - IF behavior is PARTIALLY demonstrated: then it is Coach/Attempts and assign score 5
                                - IF behavior is NOT expressed: then it is Skill Not Delivered/Unacceptabl and assign score 0
                                - IF behavior is NOT APPLICABLE: Mark as "N/A"

                            6. REASON COLUMN REQUIREMENTS - CRITICAL:   
                                - DO NOT just restate the definition in the Reason column
                                - Provide SPECIFIC analysis of what the support rep did or failed to do
                                - Explain HOW their actions/responses align or don't align with the behavior definition
                                - Reference specific policy compliance or violations when relevant
                                - For FULL scores: Explain what the rep did well that earned the maximum score and mention the which rating criteria it falls under. From the Evaluation metrics above you can refer to the three rating criteria.
                                - For PARTIAL scores: Explain what was missing or could be improved (why not full score)
                                - For ZERO scores: Explain what specific behavior was absent or done incorrectly
                                - For N/A scores: Explain why this behavior doesn't apply to this conversation
                                - Below Format to be followed without fail:
                                    - <Criteria> - <Explanation>
                                   

                            7. OUTPUT FORMAT - Use this exact table structure:

                                | Behavior | Definition | Max Score | Awarded Score | Citation | Reason |
                                    |----------|------------|-----------|---------------|----------|-----------|
                                | [Exact name] | [Definition] | ALways 10 | [Your score] | [Reference to chat history] | [Reason] |

                            8. DO NOT:
                                - Group behaviors together
                                - Create new behavior names
                                - Invent evidence not in chat history
                                - Change the behavior names provided



                        """

                    # Call LLM for evaluation
                    try:
                        response = client.chat.completions.create(
                            model="gpt-4.1",
                            messages=[
                                {
                                    "role": "system",
                                    "content": "You are an expert evaluator for customer support conversations.",
                                },
                                {"role": "user", "content": eval_prompt},
                            ],
                            max_tokens=8000,  # Increased from 1200 to handle longer evaluation tables
                            temperature=0.0,
                            top_p=1.0,  # Changed from 0.5 to 1.0 for deterministic output
                            seed=42,  # Add fixed seed for reproducibility
                            frequency_penalty=0.0,
                        )
                        eval_result = response.choices[0].message.content.strip()

                    except Exception as e:
                        eval_result = f"Error during evaluation: {str(e)}"

                    def parse_llm_table(text):
                        # Find all lines that start with '|'
                        lines = [
                            l
                            for l in text.strip().split("\n")
                            if l.strip().startswith("|")
                        ]
                        # Find the header line
                        header_idx = None
                        for i, line in enumerate(lines):
                            if "Behavior" in line and "Definition" in line:
                                header_idx = i
                                break
                        if header_idx is None or len(lines) < header_idx + 3:
                            return []
                        # Data lines start after header and separator
                        data_lines = lines[header_idx + 2 :]
                        result = []
                        for line in data_lines:
                            parts = [p.strip() for p in line.split("|")[1:-1]]
                            # Always expect 6 columns, fill missing with 'N/A'
                            while len(parts) < 6:
                                parts.append("N/A")
                            (
                                behavior,
                                definition,
                                max_score,
                                awarded_score,
                                citation,
                                reason,
                            ) = parts
                            try:
                                awarded_score_val = float(awarded_score)
                                max_score_val = float(max_score)
                                result.append(
                                    {
                                        "Behavior": behavior,
                                        "Definition": definition,
                                        "MaxScore": max_score_val,
                                        "AwardedScore": awarded_score_val,
                                        "Citation": citation,
                                        "Reason": reason,
                                    }
                                )
                            except Exception:
                                continue
                        return result

                    parsed_scores = parse_llm_table(eval_result)

                    # Only include rows where AwardedScore is a number (exclude N/A)
                    total_awarded = sum([row["AwardedScore"] for row in parsed_scores])
                    total_possible = sum([row["MaxScore"] for row in parsed_scores])
                    final_percentage = (
                        (total_awarded / total_possible * 100) if total_possible else 0
                    )

                    # SHOW SCORING FIRST
                    st.markdown("#### Evaluation Results")
                    st.markdown("---")
                    st.markdown(
                        f"**Final Score:**\n- Total Awarded: {total_awarded:.2f}\n- Total Possible: {total_possible:.2f}\n- **Final Percentage: {final_percentage:.2f}%**"
                    )

                    # THEN SHOW DETAILED TABLE
                    st.markdown("---")
                    st.markdown("#### Detailed Evaluation Table")
                    st.markdown(eval_result)

                    # --- Export to Excel Button ---

                    def export_evaluation_excel(
                        chat_history,
                        eval_table,
                        total_awarded,
                        total_possible,
                        final_percentage,
                    ):
                        wb = openpyxl.Workbook()
                        # Sheet 1: Chat History
                        ws1 = wb.active
                        ws1.title = "Chat History"
                        ws1.append(["Role", "Message"])
                        for msg in chat_history:
                            if hasattr(msg, "role"):
                                role = msg.role
                                content = msg.content
                            else:
                                role = (
                                    "user"
                                    if msg.__class__.__name__ == "HumanMessage"
                                    else "assistant"
                                )
                                content = msg.content
                            ws1.append([role, content])

                        # Sheet 2: Evaluation Table
                        ws2 = wb.create_sheet(title="Evaluation Table")
                        ws2.append(
                            [
                                "Behavior",
                                "Definition",
                                "Max Score",
                                "Awarded Score",
                                "Citation",
                                "Reason",
                            ]
                        )
                        # Use parsed_scores to write rows, always including citation
                        for row in eval_table:
                            ws2.append(
                                [
                                    row.get("Behavior", ""),
                                    row.get("Definition", ""),
                                    row.get("MaxScore", ""),
                                    row.get("AwardedScore", ""),
                                    row.get("Citation", "N/A"),
                                    row.get("Reason", ""),
                                ]
                            )

                        # Sheet 3: Final Scoring
                        ws3 = wb.create_sheet(title="Final Scoring")
                        ws3.append(
                            ["Total Awarded", "Total Possible", "Final Percentage"]
                        )
                        ws3.append(
                            [total_awarded, total_possible, f"{final_percentage:.2f}%"]
                        )

                        # NEW: Sheet 4: Metadata
                        ws4 = wb.create_sheet(title="Metadata")
                        ws4.append(
                            [
                                "username",
                                "department",
                                "issue_type",
                                "tone",
                                "timestamp",
                                "eval_score",
                            ]
                        )
                        ws4.append(
                            [
                                st.session_state.get("username", ""),
                                str(st.session_state.get("scenario_type", "")),
                                str(st.session_state.get("issue_type", "")),
                                str(st.session_state.get("customer_tone", "")),
                                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                f"{final_percentage:.2f}",
                            ]
                        )

                        output = BytesIO()
                        wb.save(output)
                        output.seek(0)
                        return output

                    excel_output = export_evaluation_excel(
                        chat_history,
                        parsed_scores,
                        total_awarded,
                        total_possible,
                        final_percentage,
                    )

                    # 🚀 OPTION 2: IMMEDIATE SAVE TO GCS (No session state dependency)
                    username = st.session_state.get("username", "")
                    if username:
                        # Get GCS storage manager
                        gcs_manager = get_gcs_storage_manager()

                        if gcs_manager.is_connected():
                            # Prepare evaluation data for immediate GCS save
                            eval_data = {
                                "username": username,
                                "timestamp": datetime.now().strftime(
                                    "%Y-%m-%d %H:%M:%S"
                                ),
                                "excel_bytes": excel_output.getvalue(),
                                "department": str(selected_domain),
                                "issue_type": str(
                                    st.session_state.get("issue_type", "")
                                ),
                                "tone": str(st.session_state.get("customer_tone", "")),
                                "eval_score": f"{final_percentage:.2f}%",
                            }

                            # Save immediately to GCS
                            success, message = gcs_manager.save_evaluation_to_gcs(
                                eval_data
                            )

                            if success:
                                st.success(f"✅ Evaluation saved to cloud storage")
                                print(
                                    f"[DEBUG] ✅ Immediate save successful: {eval_data['timestamp']}"
                                )
                                # Increment sessions_completed counter for this user's assigned simulation
                                try:
                                    assigned_sim = st.session_state.get(
                                        "assigned_simulation_dropdown"
                                    ) or st.session_state.get("assigned_simulation")
                                    username = st.session_state.get("username", "")
                                    print(
                                        f"[DEBUG] Attempting to mark completed: username='{username}', sim='{assigned_sim}'"
                                    )
                                    if assigned_sim and username:
                                        result = (
                                            user_sim_manager.mark_assignment_completed(
                                                username, assigned_sim
                                            )
                                        )
                                        print(
                                            f"[DEBUG] Mark completed result: {result}"
                                        )
                                    else:
                                        print(
                                            f"[WARN] Missing data - username: {username}, assigned_sim: {assigned_sim}"
                                        )
                                except Exception as e:
                                    # non-fatal: evaluation saved but metadata update failed
                                    print(
                                        f"[ERROR] Failed to mark assignment completed: {e}"
                                    )
                                    import traceback

                                    traceback.print_exc()
                            else:
                                st.error(f"❌ Failed to save evaluation: {message}")
                                print(f"[ERROR] ❌ Immediate save failed: {message}")
                        else:
                            st.warning(
                                "⚠️ Cloud storage unavailable. Evaluation completed but not saved."
                            )
                            print("[WARNING] GCS not connected, evaluation not saved")
                    else:
                        st.warning(
                            "⚠️ No username found. Evaluation completed but not saved."
                        )
                        print("[WARNING] No username found, evaluation not saved")


if __name__ == "__main__":
    # Ensure any startup/runtime exceptions are surfaced in logs with full tracebacks
    import traceback, sys

    try:
        main()
    except Exception as e:
        print("\n===== UNCAUGHT EXCEPTION IN app.py =====")
        print(f"Type: {type(e).__name__}")
        print(f"Error: {e}")
        traceback.print_exc()
        sys.stdout.flush()
        sys.stderr.flush()
        raise
